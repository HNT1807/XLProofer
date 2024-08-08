import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import zipfile
import io
import re


def get_column_letter(column_number):
    """Convert a column number to a column letter."""
    dividend = column_number
    column_letter = ''
    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_letter = chr(65 + modulo) + column_letter
        dividend = (dividend - modulo) // 26
    return column_letter

def check_excel_file(file):
    results = {}
    df = pd.read_excel(file)

    # Check Column Header Names
    expected_headers = [
        'RecID', 'Filename', 'Description', 'BWDescription', 'Source', 'Volume', 'Duration', 'Keywords',
        'Notes', 'Library', 'LongID', 'SampleRate', 'ShortID', 'SubCategory', 'Category', 'CDTitle',
        'CDDescription', 'TrackTitle', 'Version', 'Version_Grouping', 'Parent', 'Child', 'Composer',
        'FeaturedInstrument', 'Instrumentation', 'BPM', 'Publisher', 'Manufacturer', 'ReleaseDate',
        'Disk', 'Track', 'Show', 'Tempo', 'Location', 'Lyrics', 'TrackYear', 'LabelCode', 'Mood',
        'Usage', 'Era', 'ISWC', 'ISRC', 'ASCAPWorkno', 'BMIWorkno', 'SESACWorkno', 'GEMAWorkno',
        'Tunecode', 'HFAWorkno', 'RegistrationDate', 'ISOCountry', 'SongCode', 'SubCode', 'SongTitle',
        'SongComposer', 'Vocal', 'VocalType', 'Artist', 'PopularReference', 'FormerlyKnownAs',
        'Arranger1_First_Name', 'Arranger1_Middle_Name', 'Arranger1_Last_Name', 'Arranger1_Affiliation',
        'Arranger1_Share', 'Arranger1_CAE', 'Arranger2_First_Name', 'Arranger2_Middle_Name',
        'Arranger2_Last_Name', 'Arranger2_Affiliation', 'Arranger2_Share', 'Arranger2_CAE',
        'Composer1_First_Name', 'Composer1_Middle_Name', 'Composer1_Last_Name', 'Composer1_Affiliation',
        'Composer1_Share', 'Composer1_CAE', 'Publisher1_Name', 'Publisher1_Affiliation', 'Publisher1_Share',
        'Publisher1_CAE', 'Composer2_First_Name', 'Composer2_Middle_Name', 'Composer2_Last_Name',
        'Composer2_Affiliation', 'Composer2_Share', 'Composer2_CAE', 'Publisher2_Name',
        'Publisher2_Affiliation', 'Publisher2_Share', 'Publisher2_CAE', 'Composer3_First_Name',
        'Composer3_Middle_Name', 'Composer3_Last_Name', 'Composer3_Affiliation', 'Composer3_Share',
        'Composer3_CAE', 'Publisher3_Name', 'Publisher3_Affiliation', 'Publisher3_Share', 'Publisher3_CAE',
        'Composer4_First_Name', 'Composer4_Middle_Name', 'Composer4_Last_Name', 'Composer4_Affiliation',
        'Composer4_Share', 'Composer4_CAE', 'Publisher4_Name', 'Publisher4_Affiliation', 'Publisher4_Share',
        'Publisher4_CAE', 'Composer5_First_Name', 'Composer5_Middle_Name', 'Composer5_Last_Name',
        'Composer5_Affiliation', 'Composer5_Share', 'Composer5_CAE', 'Publisher5_Name',
        'Publisher5_Affiliation', 'Publisher5_Share', 'Publisher5_CAE', 'Composer6_First_Name',
        'Composer6_Middle_Name', 'Composer6_Last_Name', 'Composer6_Affiliation', 'Composer6_Share',
        'Composer6_CAE', 'Publisher6_Name', 'Publisher6_Affiliation', 'Publisher6_Share', 'Publisher6_CAE',
        'Composer7_First_Name', 'Composer7_Middle_Name', 'Composer7_Last_Name', 'Composer7_Affiliation',
        'Composer7_Share', 'Composer7_CAE', 'Publisher7_Name', 'Publisher7_Affiliation', 'Publisher7_Share',
        'Publisher7_CAE', 'Composer8_First_Name', 'Composer8_Middle_Name', 'Composer8_Last_Name',
        'Composer8_Affiliation', 'Composer8_Share', 'Composer8_CAE', 'Publisher8_Name',
        'Publisher8_Affiliation', 'Publisher8_Share', 'Publisher8_CAE', 'Composer9_First_Name',
        'Composer9_Middle_Name', 'Composer9_Last_Name', 'Composer9_Affiliation', 'Composer9_Share',
        'Composer9_CAE', 'Publisher9_Name', 'Publisher9_Affiliation', 'Publisher9_Share', 'Publisher9_CAE',
        'Composer10_First_Name', 'Composer10_Middle_Name', 'Composer10_Last_Name', 'Composer10_Affiliation',
        'Composer10_Share', 'Composer10_CAE', 'Publisher10_Name', 'Publisher10_Affiliation',
        'Publisher10_Share', 'Publisher10_CAE', 'MusicKey', 'Meter', 'HasLyrics', 'LyricSubject'
    ]

    missing_headers = [header for header in expected_headers if header not in df.columns]
    if not missing_headers:
        results['Column Header Names'] = '✅ <strong>COLUMN HEADER NAMES</strong>'
    else:
        results[
            'Column Header Names'] = f'❌ <strong>COLUMN HEADER NAMES</strong>|Missing/Issues with headers: {", ".join(missing_headers)}'

    # Check Column Headers Order
    if list(df.columns) == expected_headers:
        results['Column Headers Order'] = '✅ <strong>COLUMN HEADERS ORDER</strong>'
    else:
        misplaced_columns = [col for col, expected in zip(df.columns, expected_headers) if col != expected]
        misplaced_info = ', '.join(misplaced_columns[:5])  # Show up to 5 misplaced columns
        if len(misplaced_columns) > 5:
            misplaced_info += f", and {len(misplaced_columns) - 5} more"
        results[
            'Column Headers Order'] = f'❌ <strong>COLUMN HEADERS ORDER</strong>|Misplaced columns include: {misplaced_info}'


    # Check for Unwanted Spaces
    unwanted_spaces = []
    for col_num, column in enumerate(df.columns, start=1):
        for row_num, value in enumerate(df[column], start=2):
            if isinstance(value, str):
                cell_id = f"{get_column_letter(col_num)}{row_num}"
                if ' ,' in value:
                    unwanted_spaces.append(f"{cell_id} contains ' ,'")
                if ' .' in value:
                    unwanted_spaces.append(f"{cell_id} contains ' .'")
                if ' )' in value:
                    unwanted_spaces.append(f"{cell_id} contains a space before ')'")
                if '( ' in value:
                    unwanted_spaces.append(f"{cell_id} contains a space after '('")
                if ' ]' in value:
                    unwanted_spaces.append(f"{cell_id} contains a space before ']'")
                if ' |' in value:
                    unwanted_spaces.append(f"{cell_id} contains a space before '|'".replace('|', '&#124;'))
                if '| ' in value:
                    unwanted_spaces.append(f"{cell_id} contains a space after '|'".replace('|', '&#124;'))
                if '[ ' in value:
                    unwanted_spaces.append(f"{cell_id} contains a space after '['")

                if '  ' in value:
                    unwanted_spaces.append(f"{cell_id} contains a double space")

    if not unwanted_spaces:
        results['Unwanted Spaces'] = '✅ <strong>NO UNWANTED SPACES</strong>'
    else:
        results['Unwanted Spaces'] = f'❌ <strong>UNWANTED SPACES FOUND</strong>|' + '|'.join(unwanted_spaces)

    # CHECK FILENAME
    if 'Filename' in df.columns:
        invalid_filenames = []
        filename_col = df.columns.get_loc('Filename')
        filename_letter = openpyxl.utils.get_column_letter(filename_col + 1)

        def check_filename(filename, index):
            if pd.notna(filename):
                if filename.count('_') != 3:
                    invalid_filenames.append(f"{filename_letter}{index + 2} has an incorrect filename format")
                    return f"{filename_letter}{index + 2} has an incorrect filename format"
                if not filename.lower().endswith('.wav'):
                    invalid_filenames.append(f"{filename_letter}{index + 2} is not a WAV file")
                    return f"{filename_letter}{index + 2} is not a WAV file"
            return None

        df['Filename'].apply(lambda x: check_filename(x, df.index[df['Filename'] == x][0]))

        if not invalid_filenames:
            results['Filename'] = '✅ <strong>FILENAME</strong>'
        else:
            results['Filename'] = f'❌ <strong>FILENAME</strong>|' + '|'.join(invalid_filenames)
    else:
        results['Filename'] = '❌ <strong>FILENAME COLUMN NOT FOUND</strong>'

    # Check Description
    if 'Description' in df.columns and 'Filename' in df.columns:
        description_col = df.columns.get_loc('Description')
        description_letter = openpyxl.utils.get_column_letter(description_col + 1)
        invalid_descriptions = []
        for index, row in df[['Filename', 'Description']].iterrows():
            if pd.notna(row['Filename']) and pd.notna(row['Description']):
                if 'STEM' not in row['Filename'] and 'Version -' not in row['Description']:
                    invalid_descriptions.append(f"{description_letter}{index + 2} is missing \"Version -\"")
                # Check for STEM in Filename and Description
                filename = str(row['Filename'])
                if 'stem' in filename.lower():
                    if ' STEM ' not in row['Description']:
                        invalid_descriptions.append(f'{description_letter}{index + 2} is missing "STEM"')

        if not invalid_descriptions:
            results['Description'] = '✅ <strong>DESCRIPTION</strong>'
        else:
            results['Description'] = f'❌ <strong>DESCRIPTION</strong>|' + '<br>'.join(invalid_descriptions)
    else:
        pass

    # Check Source
    if 'Source' in df.columns and 'Filename' in df.columns:
        source_col = df.columns.get_loc('Source')
        source_letter = openpyxl.utils.get_column_letter(source_col + 1)
        invalid_sources = []
        for index, row in df.iterrows():
            if pd.notna(row['Filename']) and pd.notna(row['Source']):
                filename_parts = row['Filename'].split('_')
                if len(filename_parts) >= 2:
                    filename_source = '_'.join(filename_parts[:2])
                    if row['Source'] != filename_source:
                        invalid_sources.append(
                            f"{source_letter}{index + 2} '{row['Source']}' should match '{filename_source}' (from Filename)"
                        )

        if not invalid_sources:
            results['Source'] = '✅ <strong>SOURCE</strong>'
        else:
            results['Source'] = f'❌ <strong>SOURCE</strong>|' + '|'.join(invalid_sources)
    else:
        pass

    # Check Volume
    if 'Volume' in df.columns and 'Filename' in df.columns:
        volume_col = df.columns.get_loc('Volume')
        volume_letter = openpyxl.utils.get_column_letter(volume_col + 1)
        invalid_volumes = []

        for index, row in df[['Volume', 'Filename']].iterrows():
            if pd.notna(row['Volume']) and pd.notna(row['Filename']):
                expected_volume = row['Filename'].split('_')[0]
                if row['Volume'] != expected_volume:
                    invalid_volumes.append(
                        f"{volume_letter}{index + 2} '{row['Volume']}' should be '{expected_volume}'")

        if not invalid_volumes:
            results['Volume'] = '✅ <strong>VOLUME</strong>'
        else:
            results['Volume'] = f'❌ <strong>VOLUME</strong>|' + '|'.join(invalid_volumes)
    else:
        results['Volume'] = '❌ <strong>VOLUME OR FILENAME COLUMN NOT FOUND</strong>'

    def is_valid_duration(duration):
        try:
            minutes, seconds = map(int, duration.split(':'))
            return 0 <= minutes < 60 and 0 <= seconds < 60
        except:
            return False

    # In the Duration check section:
    if 'Duration' in df.columns:
        duration_col = df.columns.get_loc('Duration')
        duration_letter = openpyxl.utils.get_column_letter(duration_col + 1)
        invalid_durations = []

        for index, duration in df['Duration'].items():
            if pd.notna(duration):
                if not is_valid_duration(str(duration)):
                    invalid_durations.append(f"{duration_letter}{index + 2} ({duration}) has an invalid duration format")

        if not invalid_durations:
            results['Duration'] = '✅ <strong>DURATION</strong>'
        else:
            results['Duration'] = f'❌ <strong>DURATION</strong>|' + '|'.join(invalid_durations)
    else:
        results['Duration'] = '❌ <strong>DURATION COLUMN NOT FOUND</strong>'

    # Check BPM, Keys, and Meter
    column_display_names = {
        'BPM': 'BPM',
        'MusicKey': 'MUSIC KEY',
        'Meter': 'METER'
    }

    for column, display_name in column_display_names.items():
        if column in df.columns:
            col_index = df.columns.get_loc(column)
            col_letter = openpyxl.utils.get_column_letter(col_index + 1)
            error_messages = []

            for index, value in df[column].items():
                cell_ref = f"{col_letter}{index + 2}"
                if pd.isna(value) or value == '':
                    error_messages.append(f"{cell_ref} is missing {display_name}")
                elif column == 'BPM':
                    try:
                        if not float(value).is_integer():
                            error_messages.append(f"{cell_ref} has a BPM with a decimal")
                    except ValueError:
                        error_messages.append(f"{cell_ref} has an invalid BPM value")

            if not error_messages:
                results[column] = f'✅ <strong>{display_name}</strong>'
            else:
                results[column] = f'❌ <strong>{display_name}</strong>|' + ' | '.join(error_messages)
        else:
            results[column] = f'❌ <strong>{display_name} COLUMN NOT FOUND</strong>'

    # Check Library
    valid_libraries = [
        "COLOR TV", "ELBROAR", "MASSIVE BASS", "PERFECT PITCH", "SOUNDPORT", "INFINI", "DISKAIRE",
        "MONT CENIS", "VORTEX", "CPM", "CPM ARCHIVE SERIES", "CPM CLASSICAL", "V-THE PRODUCTION LIBRARY",
        "ONE AIR TIME", "WCBR MUSIC", "615 PLATINUM SERIES", "ESSENTIAL ELEMENTS", "GOLD SERIES",
        "KINGSIZE", "METRO", "TRUE LIFE MUSIC", "PROMO ACCELERATOR", "SCORING STAGE",
        "SPECIAL REQUESTS", "ULTIMATE CRIME & DRAMA", "CAFÉ MOONDO", "FULL TILT", "GLORY, OATH + BLOOD",
        "GRAPHIC SOUND DESIGN", "GRAVITY", "GROOVE ADDICTS", "HELLSCAPE", "IGNITE", "MINDBENDERS",
        "REVOLUCION", "WHO DID THAT MUSIC", "615", "GARI", "NON-STOP MUSIC", "ATTITUDE",
        "NON-STOP PREMIER", "NON-STOP PRODUCER SERIES", "NAKED MUSIC", "VALO ARTISTS", "VALO LATINO",
        "XPLCT MUZIK", "XTORTION AUDIO", "ADDICTED NOISE", "BIG STUFF", "CACTUS",
        "ELEPHANT SOUND DESIGN", "ELEPHANT SOUND DESIGN - WILD SANCTUARY BIOPHONIC", "PARALUX",
        "SANTA FE & 7TH", "SOUNDS FROM ECHO DISTRICT", "STORY SCORE", "SCOREMONGERS"
    ]

    if 'Library' in df.columns:
        library_col = df.columns.get_loc('Library')
        library_letter = openpyxl.utils.get_column_letter(library_col + 1)
        invalid_libraries = []

        for index, library in df['Library'].items():
            if pd.notna(library):
                if library.strip().upper() not in [lib.upper() for lib in valid_libraries]:
                    invalid_libraries.append(f"{library_letter}{index + 2}: '{library}' is not a valid option")

        if not invalid_libraries:
            results['Library'] = '✅ <strong>LIBRARY</strong>'
        else:
            results['Library'] = f'❌ <strong>LIBRARY</strong>|' + '|'.join(invalid_libraries)
    else:
        results['Library'] = '❌ <strong>LIBRARY COLUMN NOT FOUND</strong>'

    # Check LongID
    if 'LongID' in df.columns and 'Filename' in df.columns:
        longid_col = df.columns.get_loc('LongID')
        longid_letter = openpyxl.utils.get_column_letter(longid_col + 1)
        invalid_longids = []

        for index, row in df[['LongID', 'Filename']].iterrows():
            if pd.notna(row['LongID']) and pd.notna(row['Filename']):
                expected_longid = row['Filename'].rsplit('.', 1)[0]  # Remove .wav extension
                if row['LongID'] != expected_longid:
                    invalid_longids.append(
                        f"{longid_letter}{index + 2} '{row['LongID']}' doesn't match Filename")

        if not invalid_longids:
            results['LongID'] = '✅ <strong>LONG ID</strong>'
        else:
            results['LongID'] = f'❌ <strong>LONG ID</strong>|' + '|'.join(invalid_longids)
    else:
        results['LongID'] = '❌ <strong>LONG ID OR FILENAME COLUMN NOT FOUND</strong>'

    # Check SampleRate
    if 'SampleRate' in df.columns:
        samplerate_col = df.columns.get_loc('SampleRate')
        samplerate_letter = openpyxl.utils.get_column_letter(samplerate_col + 1)
        invalid_samplerates = []

        for index, samplerate in df['SampleRate'].items():
            if pd.notna(samplerate):
                if samplerate != 48000:
                    invalid_samplerates.append(f"{samplerate_letter}{index + 2} '{samplerate}' should be '48000'")

        if not invalid_samplerates:
            results['SampleRate'] = '✅ <strong>SAMPLE RATE</strong>'
        else:
            results['SampleRate'] = f'❌ <strong>SAMPLE RATE</strong>|' + '|'.join(invalid_samplerates)
    else:
        results['SampleRate'] = '❌ <strong>SAMPLE RATE COLUMN NOT FOUND</strong>'

    # Check Category
    valid_categories = [
        "Acoustic", "Adult", "Adult Contemporary", "Ambient", "Americana", "Archive",
        "Atmospheres/Drones/Beds", "Big Band", "Blues", "Children/Kids", "Choral/Chant",
        "Christmas", "Classical/Opera", "Comedy/Cartoon", "Corporate", "Country/Western",
        "Dance/Electronic", "Documentary", "Drama", "Drums/Percussion", "Easy Listening",
        "Film/TV Styles", "Folk", "Funk", "Hip-Hop/Rap", "Jazz", "Kitsch/Retro", "Latin",
        "Marches/Ceremonial/Fanfares", "National Anthem", "Orchestral/Symphonic", "Pop",
        "Pop Rock", "R&B/Soul", "Reggae", "Religious/Gospel", "Rock", "Solo/Featured Instrument",
        "Sound Design/FX", "Special Occasions", "Sports", "Traditional Dances", "Trailers",
        "Underscore", "Well Known Themes", "World Music"
    ]


    # Check Category
    if 'Category' in df.columns:
        category_col = df.columns.get_loc('Category')
        category_letter = openpyxl.utils.get_column_letter(category_col + 1)
        invalid_categories = []
        for index, cell in df['Category'].items():
            if pd.notna(cell):
                categories = [cat.strip() for cat in str(cell).split(',')]
                for cat in categories:
                    if cat not in valid_categories:
                        invalid_categories.append(f"{category_letter}{index + 2}: '{cat}' is not a valid option")

        if not invalid_categories:
            results['Category'] = '✅ <strong>CATEGORY</strong>'
        else:
            results['Category'] = f'❌ <strong>CATEGORY</strong>|' + '<br>'.join(invalid_categories)
    else:
        results['Category'] = '❌ <strong>CATEGORY COLUMN NOT FOUND</strong>'

    # Check Sub-Category
    valid_subcategories = [
        "4th of July/Independence Day", "A Capella", "Acid", "Acoustic", "Action", "Adult",
        "Adult Contemporary", "Adventure", "Africa", "Africa > Central Africa",
        "Africa > Central Africa > Angola", "Africa > Central Africa > Central African Republic",
        "Africa > Central Africa > Chad", "Africa > Central Africa > Democratic Republic of the Congo/DCR (Zaire)",
        "Africa > Central Africa > Equatorial Guinea", "Africa > Central Africa > Guinea Bissau",
        "Africa > Central Africa > Republic of the Congo", "Africa > Eastern Africa",
        "Africa > Eastern Africa > Burundi", "Africa > Eastern Africa > Eritrea",
        "Africa > Eastern Africa > Ethiopia", "Africa > Eastern Africa > Islands",
        "Africa > Eastern Africa > Kenya", "Africa > Eastern Africa > Madagascar",
        "Africa > Eastern Africa > Malawi", "Africa > Eastern Africa > Mauritius",
        "Africa > Eastern Africa > Mozambique", "Africa > Eastern Africa > Rwanda",
        "Africa > Eastern Africa > Seychelles", "Africa > Eastern Africa > Somalia",
        "Africa > Eastern Africa > Tanzania", "Africa > Eastern Africa > Uganda",
        "Africa > Eastern Africa > Zambia", "Africa > Eastern Africa > Zimbabwe",
        "Africa > Northern Africa", "Africa > Northern Africa > Algeria",
        "Africa > Northern Africa > Libya", "Africa > Northern Africa > Morocco",
        "Africa > Northern Africa > Sudan", "Africa > Northern Africa > Tunisia",
        "Africa > Southern Africa", "Africa > Southern Africa > Botswana",
        "Africa > Southern Africa > Lesotho", "Africa > Southern Africa > Namibia",
        "Africa > Southern Africa > South Africa", "Africa > Southern Africa > Swaziland",
        "Africa > Southern Africa > Zambia", "Africa > Western Africa",
        "Africa > Western Africa > Benin", "Africa > Western Africa > Burkina Faso",
        "Africa > Western Africa > Cameroon", "Africa > Western Africa > Cape Verde",
        "Africa > Western Africa > Gabon", "Africa > Western Africa > Gambia",
        "Africa > Western Africa > Ghana", "Africa > Western Africa > Guinea",
        "Africa > Western Africa > Ivory Coast", "Africa > Western Africa > Liberia",
        "Africa > Western Africa > Mali", "Africa > Western Africa > Mauritania",
        "Africa > Western Africa > Niger", "Africa > Western Africa > Nigeria",
        "Africa > Western Africa > Senegal", "Africa > Western Africa > Sierra Leone",
        "Africa > Western Africa > Togo", "Afro-Punk/AfroPunk", "AfroBeat", "Airy",
        "Alternative", "Ambient", "Americana", "Anthem", "Archive", "Artsy", "Asia",
        "Asia > Central Asia", "Asia > Central Asia > Afghanistan",
        "Asia > Central Asia > Armenia", "Asia > Central Asia > Azerbaijan",
        "Asia > Central Asia > Georgia", "Asia > Central Asia > Kazakhstan",
        "Asia > Central Asia > Kyrgyzstan", "Asia > Central Asia > Tajikistan",
        "Asia > Central Asia > Turkmenistan", "Asia > Central Asia > Uzbekistan",
        "Asia > East Asia", "Asia > East Asia > China (Peoples Republic of China)",
        "Asia > East Asia > Hong Kong", "Asia > East Asia > Japan", "Asia > East Asia > Korea",
        "Asia > East Asia > Korea > Korea, North", "Asia > East Asia > Korea > Korea, South",
        "Asia > East Asia > Macao", "Asia > East Asia > Mongolia",
        "Asia > East Asia > Taiwan (Republic of China)", "Asia > East Asia > Tibet",
        "Asia > South Asia", "Asia > South Asia > Bangladesh", "Asia > South Asia > India",
        "Asia > South Asia > Nepal", "Asia > South Asia > Pakistan", "Asia > Southeastern Asia",
        "Asia > Southeastern Asia > Bali", "Asia > Southeastern Asia > Brunei",
        "Asia > Southeastern Asia > Cambodia", "Asia > Southeastern Asia > Indonesia",
        "Asia > Southeastern Asia > Java", "Asia > Southeastern Asia > Laos",
        "Asia > Southeastern Asia > Malaysia", "Asia > Southeastern Asia > Myanmar / Burma",
        "Asia > Southeastern Asia > Philippines", "Asia > Southeastern Asia > Singapore",
        "Asia > Southeastern Asia > Thailand", "Asia > Southeastern Asia > Vietnam",
        "Asia > Southeastern Asia > Vietnam > Vietnam, North",
        "Asia > Southeastern Asia > Vietnam > Vietnam, South", "Atmospheric", "Avant Garde",
        "Bachata", "Background/Elevator", "Baion", "Ballad", "Ballroom Dance",
        "Ballroom Dance > Beguine", "Ballroom Dance > Blackbottom", "Ballroom Dance > Bolero",
        "Ballroom Dance > Bossa Nova", "Ballroom Dance > Cha Cha", "Ballroom Dance > Charleston",
        "Ballroom Dance > Conga Line", "Ballroom Dance > Foxtrot", "Ballroom Dance > Habanera",
        "Ballroom Dance > Lambada", "Ballroom Dance > Mambo", "Ballroom Dance > Merengue",
        "Ballroom Dance > Military Two Step", "Ballroom Dance > Paso Doble",
        "Ballroom Dance > Polka", "Ballroom Dance > Quickstep", "Ballroom Dance > Rhumba",
        "Ballroom Dance > Salsa", "Ballroom Dance > Samba", "Ballroom Dance > Swing/Jitterbug/Jive",
        "Ballroom Dance > Tango", "Ballroom Dance > Two Step", "Ballroom Dance > Waltz",
        "Banda", "Baroque", "Baroque Pop", "Baseball", "Basketball", "Batucada", "Beats",
        "Bebop", "Birthday", "Bluegrass", "Blues", "Bohemian", "Bomba", "Boogie Woogie",
        "Bossa Nova", "Bounce", "Boxing/UFC/Wrestling", "Brass", "Brazilian",
        "Brazilian > Afoxé", "Brazilian > Arrocha", "Brazilian > Axé", "Brazilian > Baião",
        "Brazilian > Baile Funk", "Brazilian > Boi", "Brazilian > Brega", "Brazilian > Brega funk",
        "Brazilian > Carimbó", "Brazilian > Chamamé", "Brazilian > Chorinho/Choro",
        "Brazilian > Ciranda", "Brazilian > Coco", "Brazilian > Eletrobrega",
        "Brazilian > Embolada", "Brazilian > Forró", "Brazilian > Frevo", "Brazilian > Guitarrada",
        "Brazilian > Ijexá", "Brazilian > Jongo", "Brazilian > Lundu", "Brazilian > Maculele",
        "Brazilian > Manguebeat", "Brazilian > Maracatu", "Brazilian > Marchinha",
        "Brazilian > Maxixe", "Brazilian > Modinha", "Brazilian > MPB", "Brazilian > Pagode",
        "Brazilian > Pagode Baiano", "Brazilian > Partido Alto", "Brazilian > Piseiro",
        "Brazilian > Rastapé", "Brazilian > Repente", "Brazilian > Samba de Roda",
        "Brazilian > Samba Jazz", "Brazilian > Samba Rock", "Brazilian > Sertanejo",
        "Brazilian > Tropicália", "Brazilian > Vaneira/Vaneirão", "Brazilian > Xaxado",
        "Brazilian > Xote", "Breakbeat", "Breakdance", "Bright/Optimistic", "Brit",
        "Britpop", "Buddhism", "Bumpers", "Cabaret", "Cajun", "Carols", "Cartoon", "Celtic",
        "Cha Cha", "Chamber Pop", "Champeta", "Charanga", "Chase", "Chase/Detective/Mystery",
        "Cheesy", "Chicago", "Children", "Chillout", "Chillwave", "Chiptune", "Choral",
        "Choro", "Christian", "Christmas", "Cinematic", "Classic", "Classic Rock",
        "Classic/Orchestral", "Classical", "Classical Dance", "Classical Dance > Ballet",
        "Classical Dance > Bouree", "Classical Dance > Gavotte", "Classical Dance > Gigue",
        "Classical Dance > Mazurka", "Classical Dance > Minuet", "Classical Dance > Pavane",
        "Classical Dance > Polonaise", "Classical Dance > Waltz, Classical", "Classical Form",
        "Classical Form > Adagio", "Classical Form > Aria", "Classical Form > Cadenza",
        "Classical Form > Cantata", "Classical Form > Concerto", "Classical Form > Concerto Grosso",
        "Classical Form > Etude", "Classical Form > Fugue", "Classical Form > Gregorian Chant",
        "Classical Form > Madrigal", "Classical Form > Mass", "Classical Form > Opera",
        "Classical Form > Operetta", "Classical Form > Oratorio", "Classical Form > Overture",
        "Classical Form > Prelude", "Classical Form > Rondo", "Classical Form > Sonata",
        "Classical Form > Symphony", "Classical Form > Theme and Variation",
        "Classical/Opera Fusion/Remix", "Club", "Cocktail", "College", "Colombian", "Comedy",
        "Communication/News", "Conga", "Conscious Hip-Hop", "Contemporary", "Contemporary R & B",
        "Cool", "Corrido", "Countries/English", "Countries/French", "Countries/German",
        "Countries/Italian", "Countries/Russian", "Countries/Spanish", "Country", "Cowboy",
        "Cricket", "Crime", "Crooner", "Cuban", "Cumbia", "Dance", "Dancehall", "Danza",
        "Death Metal", "Deep Funk", "Delta", "Dembow", "Detective/Mystery", "Detective/Spy",
        "Detroit Soul", "Dirge", "Dirty South/Crunk", "Disco", "Distorted", "Ditty", "Diva",
        "Dixieland", "DIY", "Documentary", "Doo Wop", "Downtempo", "Drama", "Dramedy",
        "Dream Pop", "Drinking Song", "Drone", "Drum & Bass", "Drum Corps", "Drumkit",
        "Drumline", "Drums", "Dub", "Dubstep", "Duet", "Easter", "Easy Listening", "Eclectic",
        "EDM", "Electric", "Electro", "Electro Jazz", "Electro Swing", "Electronic",
        "Electronica", "Emo", "Epic", "Ethereal", "Etherpop", "Ethnic", "Europe",
        "Europe > Eastern Europe", "Europe > Eastern Europe > Belarus",
        "Europe > Eastern Europe > Bulgaria", "Europe > Eastern Europe > Czech Republic",
        "Europe > Eastern Europe > Hungary", "Europe > Eastern Europe > Moldavia",
        "Europe > Eastern Europe > Poland", "Europe > Eastern Europe > Romania",
        "Europe > Eastern Europe > Russia / Former USSR", "Europe > Eastern Europe > Slovakia",
        "Europe > Eastern Europe > Ukraine", "Europe > Northern Europe",
        "Europe > Northern Europe > Denmark", "Europe > Northern Europe > England",
        "Europe > Northern Europe > Estonia", "Europe > Northern Europe > Finland",
        "Europe > Northern Europe > Iceland", "Europe > Northern Europe > Ireland",
        "Europe > Northern Europe > Ireland > Northern Ireland",
        "Europe > Northern Europe > Ireland > Republic of Ireland",
        "Europe > Northern Europe > Isle Of Man", "Europe > Northern Europe > Latvia",
        "Europe > Northern Europe > Lithuania", "Europe > Northern Europe > Northern Ireland",
        "Europe > Northern Europe > Norway", "Europe > Northern Europe > Scandinavia",
        "Europe > Northern Europe > Scotland", "Europe > Northern Europe > Sweden",
        "Europe > Northern Europe > United Kingdom", "Europe > Northern Europe > Wales",
        "Europe > Southern Europe", "Europe > Southern Europe > Albania",
        "Europe > Southern Europe > Bosnia Herzegovina", "Europe > Southern Europe > Croatia",
        "Europe > Southern Europe > Greece", "Europe > Southern Europe > Italy",
        "Europe > Southern Europe > Italy > Vatican City", "Europe > Southern Europe > Macedonia",
        "Europe > Southern Europe > Malta", "Europe > Southern Europe > Montenegro",
        "Europe > Southern Europe > Portugal", "Europe > Southern Europe > San Marino",
        "Europe > Southern Europe > Serbia", "Europe > Southern Europe > Slovenia",
        "Europe > Southern Europe > Spain", "Europe > Western Europe",
        "Europe > Western Europe > Andorra", "Europe > Western Europe > Austria",
        "Europe > Western Europe > Belgium", "Europe > Western Europe > France",
        "Europe > Western Europe > Germany", "Europe > Western Europe > Germany > Bavaria",
        "Europe > Western Europe > Luxembourg", "Europe > Western Europe > Monaco",
        "Europe > Western Europe > Netherlands", "Europe > Western Europe > Switzerland",
        "Exercise/Fitness/Recreational", "Exotic", "Expansive", "Experimental", "Extreme",
        "Factual", "Fairytale", "Family", "Fandango", "Fanfare/Charge", "Fanfares", "Fantasy",
        "Film/Porn", "Flamenco", "Folk", "Folk > Calypso", "Folk > Celtic", "Folk > Gypsy Style",
        "Folk > Klezmer", "Folk > Native American", "Food/Cooking", "Football > American",
        "Football > Soccer/Futbol/Rugby", "Found Sounds", "Free Jazz", "French", "Fun",
        "Funeral/Elegy", "Funk", "Fusion/Hybrid", "Future Bass", "Future Garage", "Future RnB",
        "G-Funk", "Gambling", "Game Show", "Gangsta", "Garage", "General", "German", "Glam",
        "Glitch", "Golf", "Gospel", "Gothic", "Graduation", "Grass Roots", "Grime", "Gritty",
        "Groove", "Groovy", "Grunge", "Grupero", "Guaguanco", "Guajira", "Guaracha", "Guitar",
        "Gypsy", "Halloween", "Hanukkah", "Hard", "Hard Rock", "Hardcore", "Heartland",
        "Heavy Metal", "Hick-Hop", "Hillbilly", "Hindu", "Hip Hop", "Honky Tonk", "Horror",
        "Horse/Racing/Equestrian", "House", "Hyperpop", "Ice Hockey", "IDM", "Indie",
        "Indietronica", "Industrial", "Investigative", "Islamic", "Island", "Jazz", "Jewish",
        "Jig", "Jungle", "K-Pop", "Keys/Piano", "Kiddie", "Kitschy", "Klezmer", "Kuduro",
        "Large Ensemble", "Late Night", "Latin", "Latin America",
        "Latin America > Caribbean/West Indies", "Latin America > Caribbean/West Indies > Bahamas",
        "Latin America > Caribbean/West Indies > Barbados",
        "Latin America > Caribbean/West Indies > Bermuda",
        "Latin America > Caribbean/West Indies > Cuba",
        "Latin America > Caribbean/West Indies > Dominican Republic",
        "Latin America > Caribbean/West Indies > Haiti",
        "Latin America > Caribbean/West Indies > Jamaica",
        "Latin America > Caribbean/West Indies > Puerto Rico", "Latin America > Central America",
        "Latin America > Central America > Belize", "Latin America > Central America > Costa Rica",
        "Latin America > Central America > El Salvador",
        "Latin America > Central America > Guatemala",
        "Latin America > Central America > Honduras",
        "Latin America > Central America > Mexico",
        "Latin America > Central America > Nicaragua",
        "Latin America > Central America > Panama",
        "Latin America > South America",
        "Latin America > South America > Argentina",
        "Latin America > South America > Bolivia",
        "Latin America > South America > Brazil",
        "Latin America > South America > Chile",
        "Latin America > South America > Colombia",
        "Latin America > South America > Ecuador",
        "Latin America > South America > French Guyana",
        "Latin America > South America > Guyana",
        "Latin America > South America > Paraguay",
        "Latin America > South America > Peru",
        "Latin America > South America > Suriname",
        "Latin America > South America > Uruguay",
        "Latin America > South America > Venezuela",
        "LGBTQ", "Lifestyle", "Light", "Live", "LoFi", "Lounge", "Love Song", "Lovers Rock",
        "Lullaby", "Mambo", "Mandopop", "March", "Marching Band", "Mariachi", "Martial Arts",
        "Medieval", "Mediterranean", "Merengue", "Metal", "Mexican Banda", "Miami Sound",
        "Middle East", "Middle East > Bahrain", "Middle East > Cyprus", "Middle East > Egypt",
        "Middle East > Iran", "Middle East > Iraq", "Middle East > Israel", "Middle East > Jordan",
        "Middle East > Kuwait", "Middle East > Lebanon", "Middle East > Oman", "Middle East > Qatar",
        "Middle East > Saudi Arabia", "Middle East > Syria", "Middle East > Turkey",
        "Middle East > United Arab Emirates", "Middle East > Yemen", "Midwest", "Military/War",
        "Minimal Techno", "Minimalist Style", "Modern", "Modern Blues",
        "Modern Classical/Neo Classical", "Motivational", "Motor Sports", "Motown", "Mythical",
        "Nature/Science", "Nautical", "Neo-Classical", "Neo-Soul", "New Age", "New Orleans",
        "New Wave", "New Year's Day", "News", "Newsreel", "Noir", "Norteño", "North America",
        "North America > Canada", "North America > USA",
        "North America > USA > Alaska/Pacific Northwest", "North America > USA > East Coast",
        "North America > USA > Hawaii", "North America > USA > South",
        "North America > USA > Southwest", "Nu Disco", "Nu-Folk/Pop", "Nu-Metal",
        "Nursery Rhymes/Well Known Themes", "Oceania/South Pacific",
        "Oceania/South Pacific > Australia", "Oceania/South Pacific > Cook Islands",
        "Oceania/South Pacific > Fiji", "Oceania/South Pacific > New Zealand",
        "Oceania/South Pacific > Samoa", "Oceania/South Pacific > Tahiti",
        "Oceania/South Pacific > Tonga", "Old School", "Old World", "Orchestral",
        "Orchestral/Symphonic", "Other", "Other Dance", "Other Dance > Can Can",
        "Other Dance > Go Go", "Other Dance > Jerk", "Other Dance > Soft Shoe",
        "Other Dance > Tap", "Other Dance > Twist", "Other Forms", "Other Forms > Arrangement",
        "Other Forms > Drone", "Other Forms > Fanfare", "Other Forms > Hymn", "Other Forms > Logo",
        "Other Forms > Loop/Riff", "Other Forms > Lullaby", "Other Forms > Remix",
        "Other Forms > Rhythm Track", "Pacific Northwest", "Pastoral", "Percolating", "Percussion",
        "Podcast", "Pop", "Pop Punk", "Pop Rock", "Porro", "Post", "Post Rock", "Power Ballad",
        "Power/Energetic", "Prestige/Luxury", "Prog Rock", "Progressive", "Promo", "Psychedelic",
        "Public Domain", "Pulsing", "Punk", "Quiet Storm", "Quirky", "R&B", "Ragtime", "Raï",
        "Ranchera", "Rap", "Rare Grooves", "Rave", "Raw", "Reggae", "Reggaeton", "Religious",
        "Remix", "Retro", "Rhythm & Blues", "Rhythm Bed", "Riser", "Rises", "Roadhouse", "Rock",
        "Rock n Roll", "Rockabilly", "Rocksteady", "Romantic", "Romantic Comedy", "Roots", "Royal",
        "Rural", "Salsa", "Samba", "Schlager", "Sci-Fi", "Science/Technology", "Score",
        "Second Line", "Shoegaze", "Shuffle", "Silent", "Singer Songwriter", "Sitcom", "Ska",
        "Skating/Figure Skating/Ice Skating", "Skating/Speed Skating", "Skiing", "Slapstick",
        "Small Ensemble", "Small Group", "Smooth", "Soap Opera/Telenovela", "Soca", "Soft",
        "Soft Rock", "Solea", "Son", "Song", "Soul", "Sound FX", "Soundscape", "Southern",
        "Spaghetti Western", "Spanish", "Speed Metal", "Spiritual", "Spoof", "Sports",
        "St Patrick's Day", "Stadium", "Stage Musical", "Stately", "Stings", "Stomp & Clap",
        "Storytelling", "String", "Striptease", "Surf", "Swamp", "Sweeping", "Swimming", "Swing",
        "Swooshes", "Synth", "Synthwave", "Talk Show", "Tango", "Techno", "Tejano", "Tennis",
        "Tex Mex", "Textural", "Thanksgiving", "Thrash Metal", "Thriller/Suspense",
        "Track & Field", "Traditional", "Traditional Folk/Ethnic Dance",
        "Traditional Folk/Ethnic Dance > Hora", "Traditional Folk/Ethnic Dance > Line Dance",
        "Traditional Folk/Ethnic Dance > Square Dance / Hoe Down", "Trailers /Score", "Training",
        "Trance", "Transition", "Trap", "Travel/Vacation", "Tribal", "Trio", "Trip Hop",
        "Tropical", "Tumba", "Turkish", "Tween", "Underground", "Underscore", "Valentine's Day",
        "Vallenato", "Vaudeville/Vintage", "Venezuelan > Gaita", "Venezuelan > Joropo",
        "Venezuelan > Llanera", "Vocal", "Waltz", "Weather", "Wedding", "Wellness/Relaxation",
        "West Coast", "Western", "Wind", "World", "World Games", "World Games/Sport", "Youthful",
        "Zydeco"
    ]

    if 'SubCategory' in df.columns:
        subcategory_col = df.columns.get_loc('SubCategory')
        subcategory_letter = openpyxl.utils.get_column_letter(subcategory_col + 1)
        invalid_subcategories = []
        for index, cell in df['SubCategory'].items():
            if pd.notna(cell):
                subcategories = [subcat.strip() for subcat in str(cell).split(',')]
                for subcat in subcategories:
                    if subcat not in valid_subcategories:
                        invalid_subcategories.append(
                            f"{subcategory_letter}{index + 2}: '{subcat}' is not a valid option")

        if not invalid_subcategories:
            results['SubCategory'] = '✅ <strong>SUB-CATEGORY</strong>'
        else:
            results['SubCategory'] = f'❌ <strong>SUB-CATEGORY</strong>|' + '<br>'.join(invalid_subcategories)
    else:
        results['SubCategory'] = '❌ <strong>SUB-CATEGORY COLUMN NOT FOUND</strong>'
    # Check CDTitle (assuming it should not be empty)
    if 'CDTitle' in df.columns:
        cdtitle_col = df.columns.get_loc('CDTitle')
        cdtitle_letter = openpyxl.utils.get_column_letter(cdtitle_col + 1)
        invalid_cdtitles = []
        for index, cell in df['CDTitle'].items():
            if pd.isna(cell) or str(cell).strip() == '':
                invalid_cdtitles.append(f"{cdtitle_letter}{index + 2}: CDTitle is empty")

        if not invalid_cdtitles:
            results['CDTitle'] = '✅ <strong>CD TITLE</strong>'
        else:
            results['CDTitle'] = f'❌ <strong>CD TITLE</strong>|' + '|'.join(invalid_cdtitles)
    else:
        results['CDTitle'] = '❌ <strong>CD TITLE COLUMN NOT FOUND</strong>'
    # Check TrackTitle
    if 'TrackTitle' in df.columns and 'Filename' in df.columns:
        tracktitle_col = df.columns.get_loc('TrackTitle')
        tracktitle_letter = openpyxl.utils.get_column_letter(tracktitle_col + 1)
        invalid_tracktitles = []

        for index, row in df[['TrackTitle', 'Filename']].iterrows():
            if pd.notna(row['TrackTitle']) and pd.notna(row['Filename']):
                filename_parts = row['Filename'].split('_')
                if len(filename_parts) >= 4:
                    expected_tracktitle = filename_parts[2]
                    if row['TrackTitle'] != expected_tracktitle:
                        invalid_tracktitles.append(
                            f"{tracktitle_letter}{index + 2} '{row['TrackTitle']}' should be '{expected_tracktitle}'")

        if not invalid_tracktitles:
            results['TrackTitle'] = '✅ <strong>TRACK TITLE</strong>'
        else:
            results['TrackTitle'] = f'❌ <strong>TRACK TITLE</strong>|' + '|'.join(invalid_tracktitles)
    else:
        results['TrackTitle'] = '❌ <strong>TRACK TITLE OR FILENAME COLUMN NOT FOUND</strong>'

    # Check Version
    if 'Version' in df.columns and 'Filename' in df.columns:
        version_col = df.columns.get_loc('Version')
        version_letter = openpyxl.utils.get_column_letter(version_col + 1)
        invalid_versions = []

        for index, row in df[['Version', 'Filename']].iterrows():
            if pd.notna(row['Version']) and pd.notna(row['Filename']):
                # Check for Full version
                if '_Full' in row['Filename'] and 'Full' not in row['Version']:
                    invalid_versions.append(f"{version_letter}{index + 2} should include 'Full'")

                # Check for Cutdown version
                sec_match = re.search(r'(\d+)sec', row['Filename'])
                if sec_match:
                    expected_cutdown = f"{sec_match.group(1)} Second"
                    if expected_cutdown not in row['Version']:
                        invalid_versions.append(f"{version_letter}{index + 2} should include '{expected_cutdown}'")

        if not invalid_versions:
            results['Version'] = '✅ <strong>VERSION</strong>'
        else:
            results['Version'] = f'❌ <strong>VERSION</strong>|' + '|'.join(invalid_versions)
    else:
        results['Version'] = '❌ <strong>VERSION OR FILENAME COLUMN NOT FOUND</strong>'

    # Check Version_Grouping
    if 'Version_Grouping' in df.columns and 'Version' in df.columns and 'Vocal' in df.columns and 'Filename' in df.columns:
        version_grouping_col = df.columns.get_loc('Version_Grouping')
        version_grouping_letter = openpyxl.utils.get_column_letter(version_grouping_col + 1)
        invalid_version_groupings = []

        for index, row in df[['Version_Grouping', 'Version', 'Vocal', 'Filename']].iterrows():
            if pd.notna(row['Version_Grouping']) and pd.notna(row['Version']) and pd.notna(row['Vocal']) and pd.notna(
                    row['Filename']):
                expected_version_grouping = []

                # Check for Full/Submix/Alternate
                if 'Full' in row['Version']:
                    expected_version_grouping.append('Full')
                elif 'ALT' in row['Version']:
                    expected_version_grouping.append('Alternate')
                else:
                    expected_version_grouping.append('Submix')

                # Check for Vocals/No Vocals
                if row['Vocal'] == '1':
                    expected_version_grouping.append('Vocals')
                elif row['Vocal'] == '0':
                    expected_version_grouping.append('No Vocals')

                # Check for Song
                if 'Full' in row['Version'] and ('Female' in str(row['Vocal']) or 'Male' in str(row['Vocal'])):
                    expected_version_grouping.append('Song')

                # Check for Cutdown
                sec_match = re.search(r'(\d+)sec', row['Filename'])
                if sec_match:
                    expected_version_grouping.append(f"{sec_match.group(1)} Cutdown")

                expected_version_grouping_str = ', '.join(expected_version_grouping)
                if not all(item in row['Version_Grouping'] for item in expected_version_grouping):
                    invalid_version_groupings.append(
                        f"{version_grouping_letter}{index + 2} should include '{expected_version_grouping_str}'")

        if not invalid_version_groupings:
            results['Version_Grouping'] = '✅ <strong>VERSION GROUPING</strong>'
        else:
            results['Version_Grouping'] = f'❌ <strong>VERSION GROUPING</strong>|' + '|'.join(invalid_version_groupings)
    else:
        results[
            'Version_Grouping'] = '❌ <strong>VERSION_GROUPING, VERSION, VOCAL, OR FILENAME COLUMN NOT FOUND</strong>'

    def format_track(track):
        """Preserve the original format of the track number, including leading zeros."""
        track_str = str(track)
        return track_str.zfill(len(track_str))  # Preserve original length, including leading zeros

    def get_max_track_digits(df):
        """Determine the maximum number of digits in the Track column."""
        return max(len(str(track)) for track in df['Track'] if pd.notna(track))

    def format_track(track, max_digits):
        """Format the track number with leading zeros based on the maximum number of digits."""
        return str(track).zfill(max_digits)

    def get_max_track_digits(df):
        """Determine the maximum number of digits in the Track column."""
        return max(len(str(track)) for track in df['Track'] if pd.notna(track))

    def format_track(track, max_digits):
        """Format the track number with leading zeros based on the maximum number of digits."""
        return str(track).zfill(max_digits)

    def get_max_track_digits(df):
        """Determine the maximum number of digits in the Track column."""
        return max(len(str(track)) for track in df['Track'] if pd.notna(track))

    def format_track(track, max_digits):
        """Format the track number with leading zeros based on the maximum number of digits."""
        return str(track).zfill(max_digits)

    def get_max_track_digits(df):
        """Determine the maximum number of digits in the Track column."""
        return max(len(str(track)) for track in df['Track'] if pd.notna(track))

    # Check Parent, Child, Version, Track, and TrackTitle
    if all(col in df.columns for col in ['Parent', 'Child', 'Version', 'Track', 'TrackTitle']):
        parent_col = df.columns.get_loc('Parent')
        child_col = df.columns.get_loc('Child')
        parent_letter = openpyxl.utils.get_column_letter(parent_col + 1)
        child_letter = openpyxl.utils.get_column_letter(child_col + 1)
        invalid_parent_child = []

        # Create a dictionary to store first occurrences of track titles
        track_title_first_occurrence = {}

        # First pass: record the first occurrence of each track title
        for index, row in df.iterrows():
            if pd.notna(row['TrackTitle']) and pd.notna(row['Track']):
                if row['TrackTitle'] not in track_title_first_occurrence:
                    track_title_first_occurrence[row['TrackTitle']] = (str(row['Track']), index)
                    print(f"Stored first occurrence: {row['TrackTitle']} -> {row['Track']} at row {index + 2}")

        # Second pass: check Parent and Child
        for index, row in df.iterrows():
            if pd.notna(row['Version']):
                is_full_version = 'Full' in str(row['Version'])

                # Check Parent
                if is_full_version and str(row['Parent']).strip() != 'Y':
                    invalid_parent_child.append(
                        f"{parent_letter}{index + 2} (Parent) should be 'Y' for Full version")
                elif not is_full_version and str(row['Parent']).strip() != 'N':
                    invalid_parent_child.append(
                        f"{parent_letter}{index + 2} (Parent) should be 'N' for non-Full version")

                # Check Child
                current_child = str(row['Child']).strip()
                if is_full_version:
                    expected_child = '0'
                else:
                    first_occurrence_track, first_occurrence_index = track_title_first_occurrence[row['TrackTitle']]
                    expected_child = '0' if index == first_occurrence_index else first_occurrence_track

                print(
                    f"Checking row {index + 2}: TrackTitle={row['TrackTitle']}, Track={row['Track']}, Child={current_child}, Expected Child={expected_child}")
                if current_child != expected_child:
                    invalid_parent_child.append(
                        f"{child_letter}{index + 2} (Child) should be '{expected_child}' for '{row['TrackTitle']}'")

        if not invalid_parent_child:
            results['Parent_Child'] = '✅ <strong>PARENT AND CHILD</strong>'
        else:
            results['Parent_Child'] = f'❌ <strong>PARENT AND CHILD</strong>|' + '|'.join(invalid_parent_child)
    else:
        results['Parent_Child'] = '❌ <strong>PARENT, CHILD, VERSION, TRACK, OR TRACKTITLE COLUMN NOT FOUND</strong>'

    # Check CDTitle (assuming it should not be empty)
    if 'CDTitle' in df.columns:
        cdtitle_col = df.columns.get_loc('CDTitle')
        cdtitle_letter = openpyxl.utils.get_column_letter(cdtitle_col + 1)
        invalid_cdtitles = []
        for index, cell in df['CDTitle'].items():
            if pd.isna(cell) or str(cell).strip() == '':
                invalid_cdtitles.append(f"{cdtitle_letter}{index + 2}: CDTitle is empty")

        if not invalid_cdtitles:
            results['CDTitle'] = '✅ <strong>CD TITLE</strong>'
        else:
            results['CDTitle'] = f'❌ <strong>CD TITLE</strong>|' + '|'.join(invalid_cdtitles)
    else:
        results['CDTitle'] = '❌ <strong>CD TITLE COLUMN NOT FOUND</strong>'

    def check_composer_splits(df):
        if 'Composer' not in df.columns:
            return "❌ <strong>COMPOSER SPLITS</strong>|Composer column not found"

        invalid_splits = []
        composer_col = df.columns.get_loc('Composer')
        composer_letter = openpyxl.utils.get_column_letter(composer_col + 1)

        for index, row in df.iterrows():
            if pd.notna(row['Composer']):
                percentages = re.findall(r'(\d+)%', row['Composer'])
                if percentages:
                    total = sum(int(p) for p in percentages)
                    if total != 100:
                        invalid_splits.append(f"The sum of {composer_letter}{index + 2} is {total}%, should be 100%")
                else:
                    invalid_splits.append(f"No percentages found in {composer_letter}{index + 2}")

        if not invalid_splits:
            return "✅ <strong>COMPOSER SPLITS</strong>"
        else:
            return f"❌ <strong>COMPOSER SPLITS</strong>|" + "|".join(invalid_splits)

    # Add this to your main check_excel_file function
    results['Composer Splits'] = check_composer_splits(df)

    def check_publisher_splits(df):
        if 'Publisher' not in df.columns:
            return "❌ <strong>PUBLISHER SPLITS</strong>|Publisher column not found"

        invalid_splits = []
        publisher_col = df.columns.get_loc('Publisher')
        publisher_letter = openpyxl.utils.get_column_letter(publisher_col + 1)

        for index, row in df.iterrows():
            if pd.notna(row['Publisher']):
                percentages = re.findall(r'(\d+)%', row['Publisher'])
                if percentages:
                    total = sum(int(p) for p in percentages)
                    if total != 100:
                        invalid_splits.append(f"The sum of {publisher_letter}{index + 2} is {total}%, should be 100%")
                else:
                    invalid_splits.append(f"No percentages found in {publisher_letter}{index + 2}")

        if not invalid_splits:
            return "✅ <strong>PUBLISHER SPLITS</strong>"
        else:
            return f"❌ <strong>PUBLISHER SPLITS</strong>|" + "|".join(invalid_splits)

    # Add these to your main check_excel_file function
    results['Publisher Splits'] = check_publisher_splits(df)
    # Check Manufacturer
    if 'Manufacturer' in df.columns:
        manufacturer_col = df.columns.get_loc('Manufacturer')
        manufacturer_letter = openpyxl.utils.get_column_letter(manufacturer_col + 1)
        invalid_manufacturers = []

        expected_manufacturer = "Warner Chappell Production Music"

        for index, manufacturer in df['Manufacturer'].items():
            if pd.notna(manufacturer):
                if manufacturer.strip() != expected_manufacturer:
                    invalid_manufacturers.append(
                        f"{manufacturer_letter}{index + 2}: '{manufacturer}' is incorrect")

        if not invalid_manufacturers:
            results['Manufacturer'] = '✅ <strong>MANUFACTURER</strong>'
        else:
            results['Manufacturer'] = f'❌ <strong>MANUFACTURER</strong>|' + '|'.join(invalid_manufacturers)
    else:
        results['Manufacturer'] = '❌ <strong>MANUFACTURER COLUMN NOT FOUND</strong>'

    # Check INSTRUMENTATION
    valid_instrumentations = [
        "Accordion", "Alpenhorn/Alpine Horn", "Autoharp", "Bagpipes", "Bajo Sexto", "Balafon", "Balalaika",
        "Band - Country", "Band - Dance", "Band - Ethnic", "Band - Jam", "Band - Jazz", "Band - Marching",
        "Band - Mariachi",
        "Band - Military", "Band - Oompah", "Band - Pop", "Band - Rock", "Bandoneon", "Bandura", "Banjo",
        "Bansuri/Baanhi/Baashi/Bansi/Basari", "Bass", "Bass - Acoustic", "Bass - Distorted/Fuzz", "Bass - Electric",
        "Bass - Fretless", "Bass - Slap Bass", "Bass - Slapped", "Bass - Stringed Bass/Double Bass", "Bass - Synth",
        "Bass - Upright", "Bass Drum", "Bassoon", "Batacada", "Bawu", "Bell Tree", "Bells", "Bells - Ceramic",
        "Bells - Chimes", "Bells - Church", "Bells - Hand", "Bells - Sleigh", "Bells - Tubular", "Berimbau",
        "Big Band",
        "Bladder Pipe", "Bodhran/Frame Drum", "Bombard", "Bombard - Alto", "Bombo", "Bones", "Bongos", "Bottle",
        "Bouzouki",
        "Bow", "Brass", "Brass - Brass Section", "Brass Band", "Brass Ensemble", "Bugle", "Bullroarer/Rhombus",
        "Cabasa",
        "Calliope", "Carillon", "Castanets", "Cavaquinho", "Celeste", "Cello", "Chapman Stick", "Charango",
        "Chekere/Djabara",
        "Chimes/Tubular Bells", "Cimbalom", "Cittern", "Clarinet", "Clarinet - Bass", "Clarsach", "Claves",
        "Clavinet",
        "Coconuts", "Comb And Paper", "Concertina", "Conch Shell", "Congas", "Cor Anglais/English Horn",
        "Cornamuse", "Cornet",
        "Cornett", "Cowbell", "Crotales", "Crowth", "Crumhorn", "Cuatro", "Cuica", "Cymbals", "Da Suo",
        "Daf/Dayereh", "Dan Bau",
        "Darbouka", "Def", "Descant Fiddle", "Dhol", "Dholak", "Didgeridoo", "Dilruba", "Dizi", "Djembe",
        "Dolceola", "Double Bass",
        "Doumbek/Dumbek", "Drone", "Drum Kit", "Drum Machine/Electronic Drums", "Drum Set", "Drums",
        "Drums - Brushed",
        "Drums - Cinematic/Trailer", "Drums - Distorted", "Drums - Effected", "Drums - Electronic",
        "Drums - Goblet Drum",
        "Drums - Looped", "Drums - Marching", "Drums - Military/War", "Drums - Programmed", "Drums - Sampled",
        "Duck Call",
        "Dudak", "Dudu", "Duduk", "Duff", "Dulcimer", "Dulcitone", "Dunun", "Electronic Instruments", "Electronics",
        "Erhu",
        "Esraj", "Ethnic Plucked Instruments", "Ethnic String Instruments", "Ethnic Wind Instruments", "Fiddle",
        "Fife",
        "Finger Cymbals/Finger Bells", "Finger Snaps", "Flapamba", "Flexatone", "Flugelhorn", "Flute",
        "Flute - Bamboo",
        "Flute - Chinese Bamboo", "Flute - Irish", "Flute - Native American", "Flute - Ney", "Flute - Piccolo",
        "Folk Group",
        "Football Rattle", "Fue", "Gambang", "Gamelan", "Gemshorn", "Ghaychak", "Ghurzen", "Glockenspiel",
        "Goblet Drum/Dumbec",
        "Gong", "Gong - Chinese/Chau", "Gran Cassa", "Guiro", "Guitar - 12 String",
        "Guitar - Acoustic/Nylon String",
        "Guitar - Acoustic/Steel String", "Guitar - Baritone", "Guitar - Bottleneck/Slide", "Guitar - Cigar Box",
        "Guitar - Distorted Electric", "Guitar - Dobro", "Guitar - E-Bow", "Guitar - Echo", "Guitar - Effected",
        "Guitar - Electric",
        "Guitar - Filtered", "Guitar - Fuzz", "Guitar - Gryphon", "Guitar - Harmonics", "Guitar - Hawaiian",
        "Guitar - Jazz",
        "Guitar - Nashville High String", "Guitar - Pedal Steel", "Guitar - Rhodes", "Guitar - Slack Key",
        "Guitar - Spanish ",
        "Guitar - Steel", "Guitar - Tremolo", "Guitar - Tres", "Guitar - Wah Wah", "Guitarron", "Guqin", "Guzheng",
        "Hammered Dulcimer", "Hand Claps", "Hang Drum", "Harmonica", "Harmonica - Blues Harp", "Harmonica - Glass",
        "Harmonium",
        "Harp", "Harp - Celtic", "Harpsichord", "Hi-Hat", "Hichiriki", "Horn", "Horn - French",
        "Horns/Horn Section",
        "Hurdy Gurdy", "Jazz Trio", "Jug", "Kalimba/Sanza", "Kamancheh/Kamanche/Kamancha", "Kanun", "Kaval",
        "Kawala/Salamiya",
        "Kazoo", "Kecapi", "Keyboard", "Keyboard - Piano", "Khene", "Khlui", "Koboz", "Kokyu", "Kora", "Kortholt",
        "Koto",
        "Llamas Hooves", "Log Drum", "Lute", "Lute - Chinese/Sanxian", "Lyre", "Mallet", "Mandira", "Mandocello",
        "Mandola",
        "Mandolin", "Maracas", "Marimba", "Marimbula", "MaTou Qin/Morin Khuur/Horsehead Fiddle",
        "Mbira/African Thumb Piano",
        "Mellophone", "Mellotron", "Melodeon", "Melodica", "Mizmar", "Morin Khuur", "Mouth Harp/Jews Harp",
        "Mouth/Beat Box",
        "Mridangam", "Mukkuri/Tonkori", "Musette", "Music Box", "Musical Saw", "Ney", "Ngoni", "Non-specific",
        "Novachord", "Oboe",
        "Ocarina", "Omnichord", "Orchestra", "Orchestra - Chamber", "Orchestra - Hit", "Orchestra - Hybrid",
        "Orchestra - Large",
        "Orchestra - Small", "Orchestra - Symphony", "Orchestral", "Orchestral Percussion", "Orchestral Whip",
        "Organ",
        "Organ - Accompaniment", "Organ - Barrel ", "Organ - Bontempi", "Organ - Church", "Organ - Electric",
        "Organ - Hammond",
        "Organ - Harmonium", "Organ - Mechanical", "Organ - Portative", "Organ - Wurlitzer", "Oud", "Pads",
        "Palm Court/Salon Orchestra",
        "Pan Pipes", "Percussion", "Percussion - African", "Percussion - Afro", "Percussion - Asian",
        "Percussion - Bodhran",
        "Percussion - Body", "Percussion - Chinese", "Percussion - Electronic", "Percussion - Ensemble",
        "Percussion - Ethnic",
        "Percussion - Junk/Found Objects", "Percussion - Latin", "Percussion - Metal",
        "Percussion - Middle Eastern",
        "Percussion - Native American", "Percussion - Noise", "Percussion - Tuned", "Piano",
        "Piano - Accompaniment", "Piano - Detuned/Honky Tonk",
        "Piano - Electric", "Piano - Grand", "Piano - Player Piano/Pianola", "Piano - Prepared", "Piano - Rhodes",
        "Piano - Tack",
        "Piano - Toy", "Piano - Upright", "Pipa", "Pipes", "Pipes - Celtic", "Pipes - Hornpipe", "Pipes - Pan",
        "Polyphone", "Quena",
        "Rabbi", "Rackett", "Rainstick", "Ranat", "Ratchet", "Rebec", "Recorder", "Reed Aerophone", "Riq/Kanjira",
        "Rubab/Robab/Rabab",
        "Sackbut", "Sanshin", "Santoor", "Sarangi", "Sarod", "Saunter", "Saxophone", "Saxophone - Alto",
        "Saxophone - Baritone",
        "Saxophone - Soprano", "Saxophone - Tenor", "Saz Lute/Baglama", "Scheitholt", "Scratching",
        "SFX - Ambience", "SFX - Field Recordings",
        "SFX - Noise", "SFX (Sound Effects)", "SFX (Sound Effects) - Airplanes", "SFX (Sound Effects) - Alarm",
        "SFX (Sound Effects) - Animal",
        "SFX (Sound Effects) - Applause", "SFX (Sound Effects) - Automation/Machinery",
        "SFX (Sound Effects) - Baby", "SFX (Sound Effects) - Background",
        "SFX (Sound Effects) - Birdsong", "SFX (Sound Effects) - Breathing", "SFX (Sound Effects) - Bubbles",
        "SFX (Sound Effects) - Bull/Cow",
        "SFX (Sound Effects) - Buzzer", "SFX (Sound Effects) - Camera", "SFX (Sound Effects) - Car",
        "SFX (Sound Effects) - Car Horn",
        "SFX (Sound Effects) - Chickens", "SFX (Sound Effects) - Clockwork/Grandfather Clock",
        "SFX (Sound Effects) - Comedic FX",
        "SFX (Sound Effects) - Communication", "SFX (Sound Effects) - Computer Games",
        "SFX (Sound Effects) - Cops/Police",
        "SFX (Sound Effects) - Corks Popping", "SFX (Sound Effects) - Countdown", "SFX (Sound Effects) - Crowd",
        "SFX (Sound Effects) - D.I.Y.",
        "SFX (Sound Effects) - Dog", "SFX (Sound Effects) - Door", "SFX (Sound Effects) - Drill",
        "SFX (Sound Effects) - Electronic Sounds",
        "SFX (Sound Effects) - Explosions", "SFX (Sound Effects) - Fantasy/Sci Fi Sounds",
        "SFX (Sound Effects) - Filters", "SFX (Sound Effects) - Finger Clicks",
        "SFX (Sound Effects) - Foot Stomps", "SFX (Sound Effects) - Footsteps", "SFX (Sound Effects) - Gargle",
        "SFX (Sound Effects) - Guns",
        "SFX (Sound Effects) - Hammer", "SFX (Sound Effects) - Hand Claps", "SFX (Sound Effects) - Heartbeat",
        "SFX (Sound Effects) - Helicopter",
        "SFX (Sound Effects) - Hiccup", "SFX (Sound Effects) - Hiss/Crackle", "SFX (Sound Effects) - Horns",
        "SFX (Sound Effects) - Horror", "SFX (Sound Effects) - Horses",
        "SFX (Sound Effects) - Hospital", "SFX (Sound Effects) - Human", "SFX (Sound Effects) - Human Whistle",
        "SFX (Sound Effects) - Jungle", "SFX (Sound Effects) - Laboratory",
        "SFX (Sound Effects) - Lasers/Zaps", "SFX (Sound Effects) - Laughter",
        "SFX (Sound Effects) - Martial Arts/Kung Fu", "SFX (Sound Effects) - Mechanical/Transport",
        "SFX (Sound Effects) - Military", "SFX (Sound Effects) - Milkman", "SFX (Sound Effects) - Money",
        "SFX (Sound Effects) - Nature", "SFX (Sound Effects) - Office",
        "SFX (Sound Effects) - Optical Noise", "SFX (Sound Effects) - Orchestra Tune Up",
        "SFX (Sound Effects) - Party Horn/Vuvuzela", "SFX (Sound Effects) - Party Sounds",
        "SFX (Sound Effects) - Personal Hi Fi", "SFX (Sound Effects) - Pots and Pans", "SFX (Sound Effects) - Rain",
        "SFX (Sound Effects) - Raspberry", "SFX (Sound Effects) - Record Crackle",
        "SFX (Sound Effects) - Record Scratch", "SFX (Sound Effects) - Robot",
        "SFX (Sound Effects) - Rooster/Crowing", "SFX (Sound Effects) - Running Water",
        "SFX (Sound Effects) - Scream",
        "SFX (Sound Effects) - Sea", "SFX (Sound Effects) - Sigh", "SFX (Sound Effects) - Siren",
        "SFX (Sound Effects) - Snake/Serpent", "SFX (Sound Effects) - Sneeze", "SFX (Sound Effects) - Sonar",
        "SFX (Sound Effects) - Space", "SFX (Sound Effects) - Sports FX", "SFX (Sound Effects) - Static Crackle",
        "SFX (Sound Effects) - Storm", "SFX (Sound Effects) - Sweeps/Whooshes", "SFX (Sound Effects) - Sword Fight",
        "SFX (Sound Effects) - Tarzan", "SFX (Sound Effects) - Telephone", "SFX (Sound Effects) - Tennis",
        "SFX (Sound Effects) - Thunder", "SFX (Sound Effects) - Train", "SFX (Sound Effects) - Typewriter",
        "SFX (Sound Effects) - Underwater",
        "SFX (Sound Effects) - Waves", "SFX (Sound Effects) - Whales", "SFX (Sound Effects) - Wind",
        "SFX (Sound Effects) - Zings", "Shaker", "Shakuhachi", "Shamisen", "Shawm", "Shekere", "Shenai", "Sho",
        "Shou", "Side Drum",
        "Singing Bowls", "Sitar", "Sitar - Electric", "Snare Drum", "Sound Design", "Spoons", "Steel Drums",
        "Steelpan", "Sticks", "String Ensemble", "String Quartet", "String Section", "Strings",
        "Strings - Pizzicato", "Strings - Plucked", "Suling", "Suona", "Surbahar", "Surdo", "Synth - Choir",
        "Synth - Pad", "Synth - Strings", "Synthesizer", "Synthesizer - Bells",
        "Synthesizer - Moog/Arp", "Tabla", "Taiko Drum", "Talking Drum", "Tambourine", "Tambura", "Tar", "Tarabuka",
        "Temple Bell", "Temple Blocks", "Theremin", "Thunder Sheet", "Tibetan Singing Bowls",
        "Timbale", "Timpani", "Tiompan", "Tom Toms", "Toms", "Tongue Drum", "Toy Instruments", "Transverse Flute",
        "Trautonium", "Triangle", "Tromba Marina", "Trombone", "Trumpet", "Trumpet - Muted",
        "Trumpet - Piccolo", "Tuba", "Tuba - Sousaphone", "Udu", "Ukulele", "Vibraphone", "Vibraslap", "Viol",
        "Viola", "Viola Da Gamba", "Violin", "Vocals", "Vocoder", "Washboard", "Waterphone",
        "Whip", "Whisper", "Whistle", "Whistle - Irish Low", "Whistle - Penny", "Whistle - Slide",
        "Whistle - Swanee", "Whistle - Tin", "Wind Chimes", "Wood Block", "Woodwinds", "Woodwinds - Ensemble",
        "Woodwinds - Section", "Xiao", "Xylophone ", "Yangqin", "Zagat", "Zither", "Zourna/Sorna/Zurna"
    ]

    if 'Instrumentation' in df.columns:
        instrumentation_col = df.columns.get_loc('Instrumentation')
        instrumentation_letter = openpyxl.utils.get_column_letter(instrumentation_col + 1)
        invalid_instrumentations = []
        for index, cell in df['Instrumentation'].items():
            if pd.notna(cell):
                instrumentations = [instr.strip() for instr in str(cell).split(',')]
                for instr in instrumentations:
                    if instr not in valid_instrumentations:
                        invalid_instrumentations.append(f"{instrumentation_letter}{index + 2} '{instr}' is invalid")

        if not invalid_instrumentations:
            results['Instrumentation'] = '✅ <strong>INSTRUMENTATION</strong>'
        else:
            results['Instrumentation'] = f'❌ <strong>INSTRUMENTATION</strong>|' + '|'.join(invalid_instrumentations)
    else:
        results['Instrumentation'] = '❌ <strong>INSTRUMENTATION COLUMN NOT FOUND</strong>'

    # Check Disk
    if 'Disk' in df.columns and 'Library' in df.columns:
        disk_col = df.columns.get_loc('Disk')
        disk_letter = openpyxl.utils.get_column_letter(disk_col + 1)
        invalid_disks = []

        library_initials = {
            'wcbr music': 'WCBR',
            '615 platinum series': 'SFL',
            'kingsize': 'KSM',
            'metro': 'MMP',
            'true life music': 'TLM',
            'promo accelerator': 'PA',
            'scoring stage': 'SS',
            'ultimate crime & drama': 'UCD',
            'ultimate crime and drama': 'UCD',  # Adding this variant
            'full tilt': 'FT',
            'glory, oath + blood': 'GOB',
            'gravity': 'GV',
            'groove addicts': 'GA',
            'hellscape': 'HEL',
            'ignite': 'IG',
            'mindbenders': 'MB',
            'revolucion': 'NRV',
            'who did that music': 'TL',
            'attitude': 'ATUD',
            'non-stop premier': 'NSPR',
            'non-stop producer series': 'NSPS',
            'naked music': 'NAKD',
            'valo artists': 'VALO',
            'valo latino': 'VALAT',
            'xtortion audio': 'X',
            'addicted noise': 'ADNM',
            'big stuff': 'BST',
            'cactus': 'CAC',
            'elephant sound design': 'ELEP',
            'elephant sound design - wild sanctuary biophonic': 'ELEWS',
            'paralux': 'PLX',
            'santa fe & 7th': 'SF7',
            'sounds from echo district': 'SFED',
            'story score': 'STY',
            'scoremongers': 'SCM'
        }

        for index, row in df[['Disk', 'Library']].iterrows():
            if pd.notna(row['Disk']) and pd.notna(row['Library']):
                disk_value = str(row['Disk']).replace('xxx', '').strip()  # Remove 'xxx' and strip whitespace
                library_name = str(row['Library']).lower().strip()  # Convert to lowercase and strip whitespace

                # Replace '&' with 'and' for consistency
                library_name = library_name.replace('&', 'and')

                if library_name in library_initials:
                    expected_initial = library_initials[library_name]
                    if not (disk_value.upper().startswith(expected_initial) and
                            (disk_value.upper() == expected_initial or disk_value[len(expected_initial):].isdigit())):
                        invalid_disks.append(f"{disk_letter}{index + 2} '{disk_value}' should be '{expected_initial}'")
                else:
                    invalid_disks.append(
                        f"{disk_letter}{index + 2} '{disk_value}' has an unknown library: '{row['Library']}'")

        if not invalid_disks:
            results['Disk'] = '✅ <strong>DISK</strong>'
        else:
            results['Disk'] = f'❌ <strong>DISK</strong>|' + '|'.join(invalid_disks)
    else:
        results['Disk'] = '❌ <strong>DISK OR LIBRARY COLUMN NOT FOUND</strong>'

    # Check Track
    if 'Track' in df.columns:
        track_col = df.columns.get_loc('Track')
        track_letter = openpyxl.utils.get_column_letter(track_col + 1)
        invalid_tracks = []
        track_sequence = []

        for index, track in df['Track'].items():
            if pd.notna(track):
                track = int(track)  # Convert to integer for comparison
                track_sequence.append((index, track))

        # Check for duplicates and gaps
        track_counts = {}
        for index, track in track_sequence:
            if track in track_counts:
                invalid_tracks.append(f"Issue with {track_letter}{index + 2} (duplicate number)")
            else:
                track_counts[track] = index

        expected_track = 1
        for track in sorted(track_counts.keys()):
            if track != expected_track:
                index = track_counts[track]
                invalid_tracks.append(f"Issue with {track_letter}{index + 2} (gap between numbers)")
            expected_track = track + 1

        if not invalid_tracks:
            results['Track'] = '✅ <strong>TRACK</strong>'
        else:
            results['Track'] = f'❌ <strong>TRACK</strong>|' + '|'.join(invalid_tracks)
    else:
        results['Track'] = '❌ <strong>TRACK COLUMN NOT FOUND</strong>'

    # Check TEMPO
    valid_tempos = [
        "Very Slow", "Slow", "Medium", "Medium Fast", "Fast", "Very Fast", "Variable",
        "Variable/Speeds Up", "Variable/Slows Down", "No Tempo"
    ]

    if 'Tempo' in df.columns:
        tempo_col = df.columns.get_loc('Tempo')
        tempo_letter = openpyxl.utils.get_column_letter(tempo_col + 1)
        invalid_tempos = []
        for index, cell in df['Tempo'].items():
            if pd.notna(cell):
                tempos = [tempo.strip() for tempo in str(cell).split(',')]
                for tempo in tempos:
                    if tempo not in valid_tempos:
                        invalid_tempos.append(f"{tempo_letter}{index + 2} has an invalid option ('{tempo}')")

        if not invalid_tempos:
            results['Tempo'] = '✅ <strong>TEMPO</strong>'
        else:
            results['Tempo'] = f'❌ <strong>TEMPO</strong>|' + '|'.join(invalid_tempos)
    else:
        results['Tempo'] = '❌ <strong>TEMPO COLUMN NOT FOUND</strong>'

    # Check LYRICS
    if 'Lyrics' in df.columns and 'Version_Grouping' in df.columns:
        lyrics_col = df.columns.get_loc('Lyrics')
        lyrics_letter = get_column_letter(lyrics_col + 1)
        invalid_lyrics = []

        for index, row in df.iterrows():
            if pd.notna(row['Version_Grouping']) and ', Lyrics' in row['Version_Grouping']:
                if pd.isna(row['Lyrics']) or (isinstance(row['Lyrics'], str) and row['Lyrics'].strip() == ''):
                    invalid_lyrics.append(f"{lyrics_letter}{index + 2} is missing lyrics")

        if not invalid_lyrics:
            results['Lyrics'] = '✅ <strong>LYRICS</strong>'
        else:
            results['Lyrics'] = f'❌ <strong>LYRICS</strong>|' + '|'.join(invalid_lyrics)
    else:
        results['Lyrics'] = '❌ <strong>LYRICS OR VERSION_GROUPING COLUMN NOT FOUND</strong>'

    # Check TRACK YEAR
    if 'TrackYear' in df.columns:
        track_year_col = df.columns.get_loc('TrackYear')
        track_year_letter = get_column_letter(track_year_col + 1)
        invalid_track_years = []

        for index, row in df.iterrows():
            if not row.isnull().all():  # Check if the entire row is not empty
                if pd.isna(row['TrackYear']) or (
                        isinstance(row['TrackYear'], str) and row['TrackYear'].strip() == '') or (
                        isinstance(row['TrackYear'], (int, float)) and row['TrackYear'] == 0):
                    invalid_track_years.append(f"{track_year_letter}{index + 2} is missing track year")

        if not invalid_track_years:
            results['TrackYear'] = '✅ <strong>TRACK YEAR</strong>'
        else:
            results['TrackYear'] = f'❌ <strong>TRACK YEAR</strong>|' + '|'.join(invalid_track_years)
    else:
        results['TrackYear'] = '❌ <strong>TRACK YEAR COLUMN NOT FOUND</strong>'
    label_codes = {
        "1st Producer Series": "96445",
        "2nd Foundation Music": "51208",
        "2 Red Jokers": "99431",
        "615 Platinum Series": "30201",
        "Gold Series": "12540",
        "Acoustitracks": "29655",
        "Addicted Noise": "100328",
        "Alchemy Music": "29725",
        "Alchemy Trailer Tools": "89364",
        "Amphibious Zoo": "29656",
        "Anarchy": "24364",
        "Arrow Production Music": "89342",
        "Art of Legend": "85158",
        "Attitude": "30275",
        "Audioactive": "29657",
        "Audio Attack Production Music": "96527",
        "Audio Attack Trailer Series": "96528",
        "Audio Junkies": "98715",
        "Audiomachine": "29653",
        "Audiomonkey": "51627",
        "Authentricity": "98717",
        "Avalon Zero": "100179",
        "Big Stuff": "99459",
        "Cactus": "99034",
        "Cafe Moondo": "27120",
        "Cinema Sound Tools": "27122",
        "ColorTV": "52712",
        "CPM": "07189",
        "CPM Archive Series": "30199",
        "CPM Classical": "30200",
        "CrimeSonics": "97413",
        "Darwin Music": "85157",
        "Diskaire": "99201",
        "Ear Drum": "18761",
        "Elbroar": "52589",
        "Elephant Sound Design": "24367",
        "Elephant Sound Design - Wild Sanctuary Biophonic": "98790",
        "EMH Classical": "29654",
        "Enterprises Sonor Production Music": "49199",
        "Epic Single Series": "96529",
        "Essential Elements": "28036",
        "Full Tilt": "18759",
        "Future Pop": "86347",
        "Future Pop's TV Tools": "98846",
        "Glory FX": "101731",
        "Glory Oath & Blood": "19837",
        "Gothic Storm's Comedy Works": "99315",
        "Gothic Hybrid": "85033",
        "Gothic Storm Music": "24723",
        "Gothic Storm Toolworks": "85035",
        "Graphic Sound Design": "27121",
        "Gravity": "11746",
        "Grey Area Sound (GAS)": "29659",
        "Grey Area Sound (OIL)": "30206",
        "Groove Addicts": "18758",
        "Hellscape": "98270",
        "Hit Music Lab": "99194",
        "Hot Tag Media": "92612",
        "Ignite": "11745",
        "Image": "04099",
        "Image Edge": "30203",
        "Impressive Minds": "51584",
        "Imtel": "30202",
        "Indiesonics Library": "51330",
        "Infini Music": "24408",
        "Kingsize": "28099",
        "Kurt Bestor Music": "77502",
        "Library of the Human Soul": "49016",
        "Lovely Music": "51336",
        "Magnum Opus PM": "101042",
        "Massive Bass": "30095",
        "Metro": "28041",
        "Mathambo Music": "29686",
        "Marigold": "98716",
        "MidCoast Music Artists Songs": "35163",
        "MidCoast Music Special Occasions": "86434",
        "MidCoast Music Wired": "33780",
        "Mind Benders": "27119",
        "Minimal": "28037",
        "Minim Music": "57978",
        "Mont Cenis": "101509",
        "Naked Music": "34666",
        "New Revolution": "30198",
        "Non-Stop Attitude": "30196",
        "Non-Stop Premiere": "30197",
        "Non-Stop Producer Series": "08903",
        "One Air Time": "49301",
        "Paralux": "100596",
        "Perfect Pitch": "24083",
        "Popcorn Music": "86349",
        "Production TRX": "30205",
        "Promo Accelerator": "28038",
        "Revolution Music": "11748",
        "Ritual Echoes Music": "91802",
        "Run4Cover": "98612",
        "Santa Fe & Seventh": "99473",
        "Sauce Music": "24805",
        "Scaremeister": "19836",
        "Score TRX": "30204",
        "Scoremongers": "95636",
        "Scoring Stage": "28039",
        "Silent Methods": "18760",
        "Sky Urbano": "98714",
        "Songcraft": "86355",
        "Soundport": "98675",
        "Sounds From Echo District": "100035",
        "Special Request": "28042",
        "StoryScore": "98822",
        "Tenth Dimension": "29658",
        "The Fix": "100036",
        "Timber Music Supply": "98992",
        "True Life Music": "95617",
        "TRX Music Library": "13379",
        "Ultimate Crime & Drama": "28040",
        "VALO Artists": "51690",
        "VALO Latin": "91362",
        "V - The Production Library": "13333",
        "Vortex Music": "100057",
        "Warner Chappell Music Italiana": "29705",
        "WCBR Music": "101061",
        "Who Did That Music": "11747",
        "Xtortion": "29759"
    }

    # Check LABEL CODE
    if 'LabelCode' in df.columns and 'Library' in df.columns:
        label_code_col = df.columns.get_loc('LabelCode')
        label_code_letter = get_column_letter(label_code_col + 1)
        invalid_label_codes = []

        for index, row in df.iterrows():
            if pd.notna(row['Library']):
                expected_code = label_codes.get(row['Library'])
                if expected_code:
                    if pd.isna(row['LabelCode']) or str(row['LabelCode']).strip() == '':
                        invalid_label_codes.append(
                            f"{label_code_letter}{index + 2} is missing label code ({expected_code})")
                    else:
                        # Convert both expected and actual codes to integers for comparison
                        actual_code = int(float(row['LabelCode']))
                        expected_code = int(expected_code)
                        if actual_code != expected_code:
                            invalid_label_codes.append(
                                f"{label_code_letter}{index + 2} should be '{expected_code}' not '{actual_code}'")

        if not invalid_label_codes:
            results['LabelCode'] = '✅ <strong>LABEL CODE</strong>'
        else:
            results['LabelCode'] = f'❌ <strong>LABEL CODE</strong>|' + '|'.join(invalid_label_codes)
    else:
        results['LabelCode'] = '❌ <strong>LABEL CODE OR LIBRARY COLUMN NOT FOUND</strong>'

    # Check Mood
    valid_moods = [
        "Abstract", "Adventurous", "Aftermath", "Aggressive", "Angry", "Anticipation/Anticipatory",
        "Anxious/Nervous", "Artful", "Awkward", "Bittersweet", "Bold", "Bouncy", "Breezy",
        "Bright", "Building/Rising", "Calm", "Campy", "Cautious", "Celebratory", "Celestial/Ethereal",
        "Cerebral", "Chaotic", "Cheeky", "Cold", "Comical/Humorous", "Confident/Strong/Proud",
        "Cool", "Creepy", "Curious", "Dangerous/Threatening", "Dark", "Defeated", "Delicate",
        "Determined", "Dramatic", "Dreamy", "Driving", "Drunk/Stoned", "Dynamic", "Eccentric/Quirky",
        "Edgy", "Eerie", "Elegant/Graceful", "Emotional", "Energetic", "Epic", "Euphoric", "Evil",
        "Excited", "Exotic", "Floating", "Fragile", "Frenzied/Zany", "Fun", "Funky", "Happy",
        "Happy/Positive/Bright", "Hard", "Heartbroken", "Heavy", "Hopeful", "Humorous", "Hypnotic",
        "Indecisive", "Insistent", "Inspirational/Uplifting", "Intense", "Intimate", "Intrigue",
        "Jazzy", "Laidback", "Lighthearted", "Lonely", "Love/Romantic", "Macabre", "Majestic",
        "Mechanical/Robotic", "Melancholy", "Mellow", "Mischievous", "Mysterious", "Neutral",
        "Noble/Majestic/Prestigious", "Passionate", "Pathos", "Patriotic", "Peaceful", "Playful",
        "Poignant", "Positive", "Powerful", "Proud/Motivational", "Psychedelic", "Quirky",
        "Reflective", "Relaxed", "Rhythmic", "Running", "Sacred", "Sad/Depressing", "Scary",
        "Seductive", "Sensitive", "Serene", "Sexy", "Shy", "Simple", "Slow", "Slow Motion",
        "Solemn", "Spiritual", "Spooky", "Static", "Stealthy", "Suspense/Tension", "Tender",
        "Tense", "Thoughtful", "Time Passing/Time Lapse", "Tragic", "Transcendant", "Urgent",
        "Victorious/Triumphant", "Violent", "Walking", "Warm", "Warning", "Weird/Confused",
        "Whimsical", "Wistful", "Wondrous/Breathtaking"
    ]

    if 'Mood' in df.columns:
        mood_col = df.columns.get_loc('Mood')
        mood_letter = openpyxl.utils.get_column_letter(mood_col + 1)
        invalid_moods = []
        for index, cell in df['Mood'].items():
            if pd.notna(cell):
                moods = [mood.strip() for mood in str(cell).split(',')]
                for mood in moods:
                    if mood not in valid_moods:
                        invalid_moods.append(f"{mood_letter}{index + 2}: '{mood}' is not a valid option")

        if not invalid_moods:
            results['Mood'] = '✅ <strong>MOOD</strong>'
        else:
            results['Mood'] = f'❌ <strong>MOOD</strong>|' + '<br>'.join(invalid_moods)
    else:
        results['Mood'] = '❌ <strong>MOOD COLUMN NOT FOUND</strong>'

    # Check Usage
    valid_usages = [
        "Action", "Adult Content", "Adventure", "Advertising", "Animation", "Arts", "Award Show",
        "Beauty/Cosmetics", "Bed", "Business/Corporate", "Ceremony", "Children/Kids", "Christmas",
        "Circus", "Comedy", "Coming of Age", "Competition Shows", "Crime/Investigation", "Discovery",
        "DIY", "Documentary", "Drama", "Fantasy/Magical", "Fashion/Makeover", "Film Trailer", "Foley",
        "Food/Drink", "Friendship/Brotherhood/Sisterhood", "Game Show", "Historical/Period Piece",
        "Holiday/Seasonal", "Horror", "Human Interest", "Jingle", "Journey", "Melodrama", "Mystery",
        "Nature/Wildlife/The Elements", "News", "Political/Election", "Production Elements", "Promo",
        "Protest", "Reality TV", "Religion", "Reveal", "Road Trip", "Romance", "Royal Event",
        "Science/Technology/Medicine", "Sexy", "Shopping/Retail", "Soap Opera/Telenovela",
        "Special Occasions", "Sports", "Talk Show", "Tasking", "Thriller", "Travel/Vacation",
        "Video Games", "War/Conflict", "Wedding", "Western"
    ]

    if 'Usage' in df.columns:
        usage_col = df.columns.get_loc('Usage')
        usage_letter = openpyxl.utils.get_column_letter(usage_col + 1)
        invalid_usages = []
        for index, cell in df['Usage'].items():
            if pd.notna(cell):
                usages = [usage.strip() for usage in str(cell).split(',')]
                for usage in usages:
                    if usage not in valid_usages:
                        invalid_usages.append(f"{usage_letter}{index + 2}: '{usage}' is not a valid option")

        if not invalid_usages:
            results['Usage'] = '✅ <strong>USAGE</strong>'
        else:
            results['Usage'] = f'❌ <strong>USAGE</strong>|' + '<br>'.join(invalid_usages)
    else:
        results['Usage'] = '❌ <strong>USAGE COLUMN NOT FOUND</strong>'


    # Check ERA
    valid_eras = [
        "2020s/Future", "2010s", "2000s", "1990s", "1980s", "1970s", "1960s", "1950s", "1940s",
        "1930s", "1920s", "1910s", "1900s", "1830 - 1900 Romantic", "1750 - 1830 Classical",
        "1600 - 1750 Baroque", "1400 - 1600 Renaissance", "1150 - 1400 Medieval",
        "0 - 1150 Early Music/Ancient", "3100 BC - 1 BC Ancient Greek"
    ]

    if 'Era' in df.columns:
        era_col = df.columns.get_loc('Era')
        era_letter = openpyxl.utils.get_column_letter(era_col + 1)
        invalid_eras = []
        for index, cell in df['Era'].items():
            if pd.notna(cell):
                eras = [era.strip() for era in str(cell).split(',')]
                for era in eras:
                    if era not in valid_eras:
                        invalid_eras.append(f"{era_letter}{index + 2}: '{era}' is not a valid option")

        if not invalid_eras:
            results['Era'] = '✅ <strong>ERA</strong>'
        else:
            results['Era'] = f'❌ <strong>ERA</strong>|' + '<br>'.join(invalid_eras)
    else:
        results['Era'] = '❌ <strong>ERA COLUMN NOT FOUND</strong>'

    # Check VOCAL
    if 'Vocal' in df.columns and 'Version_Grouping' in df.columns:
        vocal_col = df.columns.get_loc('Vocal')
        vocal_letter = get_column_letter(vocal_col + 1)
        invalid_vocals = []

        for index, row in df[['Vocal', 'Version_Grouping']].iterrows():
            if pd.notna(row['Version_Grouping']):
                expected_vocal = '1' if ", Vocal" in row['Version_Grouping'] else '0'
                if pd.notna(row['Vocal']) and str(row['Vocal']) != expected_vocal:
                    invalid_vocals.append(f"{vocal_letter}{index + 2} should be {expected_vocal} not {row['Vocal']}")
            elif pd.notna(row['Vocal']):
                invalid_vocals.append(f"{vocal_letter}{index + 2} Vocal value present but Version_Grouping is empty")

        if not invalid_vocals:
            results['Vocal'] = '✅ <strong>VOCAL</strong>'
        else:
            results['Vocal'] = f'❌ <strong>VOCAL</strong>|' + '|'.join(invalid_vocals)
    else:
        results['Vocal'] = '❌ <strong>VOCAL OR VERSION_GROUPING COLUMN NOT FOUND</strong>'

    # Check VOCAL TYPE
    valid_vocal_types = [
        "No Vocal", "Barber Shop", "Child", "Child - Boy", "Child - Girl", "Choir", "Choir - Boy", "Choir - Female",
        "Choir - Girl", "Choir - Male", "Choir - Mixed Adult", "Choir - Mixed Adult Child", "Choir - Mixed Child",
        "Crooner", "Duet - Boy Female", "Duet - Boy Girl", "Duet - Boy Male", "Duet - Female Female",
        "Duet - Girl Female", "Duet - Girl Male", "Duet - Male Female", "Duet - Male Male", "Explicit", "Female",
        "Female - Alto/Contralto", "Female - Explicit", "Female - Female", "Female - Mezzo Soprano",
        "Female - Soprano", "Male", "Male - Baritone", "Male - Bass", "Male - Explicit", "Male - Male",
        "Male - Tenor", "Overtone Singing/Throat Singing", "Scat Singing", "Synth Voice/Vocoder", "Vocal Textures",
        "Vocal Textures - Beatbox", "Vocal Textures - Chanting Stadium", "Vocal Textures - Humming",
        "Vocal Textures - Speechless", "Vocal Textures - Vocal Background", "Vocal Textures - Vocal Phrase/Shout Out",
        "Whistling", "Yodeling - Alpine", "Yodeling - Western/USA"
    ]

    if 'VocalType' in df.columns:
        vocal_type_col = df.columns.get_loc('VocalType')
        vocal_type_letter = get_column_letter(vocal_type_col + 1)
        invalid_vocal_types = []

        for index, cell in df['VocalType'].items():
            if pd.notna(cell):
                types = [vtype.strip() for vtype in str(cell).split(',')]
                for vtype in types:
                    if vtype not in valid_vocal_types:
                        invalid_vocal_types.append(f"{vocal_type_letter}{index + 2}: '{vtype}' is not a valid option")

        if not invalid_vocal_types:
            results['VocalType'] = '✅ <strong>VOCAL TYPE</strong>'
        else:
            results['VocalType'] = f'❌ <strong>VOCAL TYPE</strong>|' + '|'.join(invalid_vocal_types)
    else:
        results['VocalType'] = '❌ <strong>VOCAL TYPE COLUMN NOT FOUND</strong>'

    def parse_composer_info(composer_string):
        composers = []
        for composer in composer_string.split(','):
            composer = composer.strip()
            name_parts = composer.split('(')[0].strip().split()
            first_name = name_parts[0]
            last_name = name_parts[-1] if len(name_parts) > 1 else ''
            pro = re.search(r'\((.*?)\)', composer)
            pro = pro.group(1) if pro else ''
            share = re.search(r'(\d+(?:\.\d+)?)', composer)
            share = share.group(1) if share else ''
            cae = re.search(r'\[(.*?)\]', composer)
            cae = cae.group(1) if cae else ''
            composers.append({
                'first_name': first_name,
                'last_name': last_name,
                'affiliation': pro,
                'share': share,
                'cae': cae
            })
        return composers

    # Check Composer Information
    if all(col in df.columns for col in ['Composer'] + [f'Composer{i}_{field}' for i in range(1, 11) for field in
                                                        ['First_Name', 'Last_Name', 'Affiliation', 'Share',
                                                         'CAE']]):
        composer_col = df.columns.get_loc('Composer')
        composer_letter = get_column_letter(composer_col + 1)
        invalid_composer_info = []

        for index, row in df.iterrows():
            if pd.notna(row['Composer']):
                parsed_composers = parse_composer_info(row['Composer'])
                for i, composer in enumerate(parsed_composers, start=1):
                    for field in ['First_Name', 'Last_Name', 'Affiliation', 'Share', 'CAE']:
                        col_name = f'Composer{i}_{field}'
                        actual_value = row[col_name]
                        expected_value = composer[field.lower()]

                        # Convert empty cells and 'nan' to 'Nothing'
                        if pd.isna(actual_value) or str(actual_value).strip().lower() == 'nan' or str(
                                actual_value).strip() == '':
                            actual_value = 'Nothing'
                        else:
                            actual_value = str(actual_value).strip()

                        expected_value = 'Nothing' if expected_value == '' else str(expected_value).strip()

                        # For Share and CAE fields, compare numeric values
                        if field in ['Share', 'CAE']:
                            try:
                                actual_numeric = float(actual_value)
                                expected_numeric = float(expected_value)
                                if abs(actual_numeric - expected_numeric) < 1e-6:  # Use a small tolerance for floating-point comparison
                                    continue
                            except ValueError:
                                pass  # If conversion fails, fall back to string comparison

                        if field == 'Last_Name' and not composer['last_name']:
                            if actual_value != 'Nothing':
                                invalid_composer_info.append(
                                    f"{get_column_letter(df.columns.get_loc(col_name) + 1)}{index + 2}: Last Name for Composer {i}. Found: '{actual_value}', Expected: Nothing")
                        elif actual_value != expected_value:
                            invalid_composer_info.append(
                                f"{get_column_letter(df.columns.get_loc(col_name) + 1)}{index + 2}: {field} for Composer {i}. Found: '{actual_value}', Expected: '{expected_value}'")

        if not invalid_composer_info:
            results['Composer Information'] = '✅ <strong>COMPOSER INFORMATION</strong>'
        else:
            results['Composer Information'] = f'❌ <strong>COMPOSER INFORMATION</strong>|' + '|'.join(
                invalid_composer_info)
    else:
        results['Composer Information'] = '❌ <strong>COMPOSER INFORMATION COLUMNS NOT FOUND</strong>'

    def parse_publisher_info(publisher_string):
        publishers = []
        for publisher in publisher_string.split(','):
            publisher = publisher.strip()
            name = publisher.split('(')[0].strip()
            affiliation = re.search(r'\((.*?)\)', publisher)
            affiliation = affiliation.group(1) if affiliation else ''
            share = re.search(r'(\d+(?:\.\d+)?)%', publisher)
            share = share.group(1) if share else ''
            cae = re.search(r'\[(.*?)\]', publisher)
            cae = cae.group(1) if cae else ''
            publishers.append({
                'name': name,
                'affiliation': affiliation,
                'share': share,
                'cae': cae
            })
        return publishers

    # Check Publisher Information
    if all(col in df.columns for col in ['Publisher'] + [f'Publisher{i}_{field}' for i in range(1, 11) for field in
                                                         ['Name', 'Affiliation', 'Share', 'CAE']]):
        publisher_col = df.columns.get_loc('Publisher')
        publisher_letter = get_column_letter(publisher_col + 1)
        invalid_publisher_info = []

        for index, row in df.iterrows():
            if pd.notna(row['Publisher']):
                parsed_publishers = parse_publisher_info(row['Publisher'])
                for i, publisher in enumerate(parsed_publishers, start=1):
                    for field in ['Name', 'Affiliation', 'Share', 'CAE']:
                        col_name = f'Publisher{i}_{field}'
                        actual_value = row[col_name]
                        expected_value = publisher[field.lower()]

                        # Convert empty cells and 'nan' to 'Nothing'
                        if pd.isna(actual_value) or str(actual_value).strip().lower() == 'nan' or str(
                                actual_value).strip() == '':
                            actual_value = 'Nothing'
                        else:
                            actual_value = str(actual_value).strip()

                        expected_value = 'Nothing' if expected_value == '' else str(expected_value).strip()

                        # For Share and CAE fields, compare numeric values
                        if field in ['Share', 'CAE']:
                            try:
                                actual_numeric = float(actual_value)
                                expected_numeric = float(expected_value)
                                if abs(actual_numeric - expected_numeric) < 1e-6:  # Use a small tolerance for floating-point comparison
                                    continue
                            except ValueError:
                                pass  # If conversion fails, fall back to string comparison

                        if actual_value != expected_value:
                            invalid_publisher_info.append(
                                f"{get_column_letter(df.columns.get_loc(col_name) + 1)}{index + 2}: {field} for Publisher {i}. Found: '{actual_value}', Expected: '{expected_value}'")

        if not invalid_publisher_info:
            results['Publisher Information'] = '✅ <strong>PUBLISHER INFORMATION</strong>'
        else:
            results['Publisher Information'] = f'❌ <strong>PUBLISHER INFORMATION</strong>|' + '|'.join(
                invalid_publisher_info)
    else:
        results['Publisher Information'] = '❌ <strong>PUBLISHER INFORMATION COLUMNS NOT FOUND</strong>'




    return results




def main():
    st.set_page_config(page_title="XL Proofer", layout="wide")

    left_column, center_column, right_column = st.columns([1, 2, 1])

    with center_column:
        st.markdown("""
                <style>
                .center-title {
                    text-align: center;
                }
                .center-content {
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                }
                </style>
                <div class="center-content">
                    <h1 class="center-title">XL PROOFER</h1>
                """, unsafe_allow_html=True)

        uploaded_files = st.file_uploader("Upload Excel files", type=['xls', 'xlsx'], accept_multiple_files=True)

        marked_down_files = {}  # Store all marked-down files

        if uploaded_files:
            for file in uploaded_files:
                with st.expander(f"Results for {file.name}"):
                    results = check_excel_file(file)

                    passed = [value for value in results.values() if value.startswith('✅')]
                    failed = [value for value in results.values() if value.startswith('❌')]

                    # Display passed checks
                    for value in passed:
                        st.markdown(f'<p class="check-item">{value}</p>', unsafe_allow_html=True)

                    # Display failed checks
                    for i, value in enumerate(failed):
                        parts = value.split('|')
                        title = parts[0]
                        details = parts[1:] if len(parts) > 1 else []

                        # Add margin to the last item
                        margin_class = ' last-item' if i == len(failed) - 1 else ''

                        details_html = '<br>'.join(details) if details else 'No additional details.'

                        st.markdown(f"""
                            <details class="check-item{margin_class}">
                                <summary>{title}</summary>
                                <div class="check-details">{details_html}</div>
                            </details>
                            """, unsafe_allow_html=True)

                    # Add marked-down file to the dictionary
                    if failed:
                        marked_down_file = create_marked_down_excel(file, results)
                        marked_down_files[f"marked_down_{file.name}"] = marked_down_file
                        st.download_button(
                            label="DOWNLOAD MARKED DOWN XL",
                            data=marked_down_file,
                            file_name=f"marked_down_{file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

        # Add "DOWNLOAD ALL MARKED DOWN XLS" button
        if marked_down_files:
            zip_buffer = create_zip_of_marked_down_files(marked_down_files)
            st.download_button(
                label="DOWNLOAD ALL MARKED DOWN XLS",
                data=zip_buffer,
                file_name="all_marked_down_files.zip",
                mime="application/zip"
            )


    # Add some custom CSS to ensure the content doesn't exceed the column width and add space after the last item
    st.markdown("""
    <style>
    .stApp > header {
        background-color: transparent;
    }
    .stApp {
        max-width: 1200px;
        margin: 0 auto;
    }
    .check-item {
        margin-bottom: 0.5rem;
    }
    .check-details {
        margin-top: 0.5rem;
    }
    .last-item {
        margin-bottom: 20px;
    }
    </style>
    """, unsafe_allow_html=True)

def create_zip_of_marked_down_files(marked_down_files):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for file_name, file_content in marked_down_files.items():
            zip_file.writestr(file_name, file_content.getvalue())
    zip_buffer.seek(0)
    return zip_buffer


def create_marked_down_excel(file, results):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')

    def highlight_specific_field(field_name):
        if field_name in results and results[field_name].startswith('❌'):
            column_letter = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == field_name:
                    column_letter = openpyxl.utils.get_column_letter(col)
                    break

            if column_letter:
                error_details = results[field_name].split('|')[1].split('<br>')
                for detail in error_details:
                    match = re.search(r'([A-Z]+)(\d+)', detail)
                    if match:
                        row = int(match.group(2))
                        try:
                            ws[f'{column_letter}{row}'].fill = orange_fill
                        except ValueError as e:
                            print(f"Error highlighting cell {column_letter}{row} for {field_name}: {str(e)}")

    # Highlight Category errors
    highlight_specific_field('Category')

    # Highlight Era errors
    highlight_specific_field('Era')

    # Highlight other fields (if needed)
    for field in ['Filename', 'Description', 'Source', 'Volume', 'Duration', 'Library',
                  'CDTitle', 'Instrumentation', 'BPM', 'Tempo', 'LongID', 'SampleRate', 'TrackTitle', 'Version',
                  'Version_Grouping', 'Parent_Child', 'Composer Splits', 'Disk', 'Manufacturer',
                  'Track', 'Lyrics', 'TrackYear', 'LabelCode', 'Vocal', 'VocalType',
                  'Composer Information', 'Publisher Information', 'SubCategory', 'Mood', 'Usage']:
        if field in results and results[field].startswith('❌'):
            details = results[field].split('|')[1:]
            for detail in details:
                cell_ids = re.findall(r'([A-Z]+\d+)', detail)
                for cell_id in cell_ids:
                    try:
                        ws[cell_id].fill = orange_fill
                    except ValueError as e:
                        print(f"Error highlighting cell {cell_id} for {field}: {str(e)}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


if __name__ == "__main__":
    main()
